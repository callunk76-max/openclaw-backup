from flask import Flask, render_template, request, send_file, jsonify
import sqlite3
import pandas as pd
import os
import json
import csv
import time
from io import BytesIO
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

app = Flask(__name__)

VISITOR_PATH = '/root/.openclaw/workspace/App/sirup_bulukumba/visitor_count.json'
DB_PATH = '/root/.openclaw/workspace/App/sirup_bulukumba/bulukumba.db'

BUDGET_SQL = "CAST(REPLACE(REPLACE(budget, 'Rp ', ''), ',', '') AS REAL)"
PER_PAGE = 50

# Simple TTL cache
_cache = {}
_CACHE_TTL = 300  # 5 minutes

def cached(key, ttl=_CACHE_TTL):
    def decorator(func):
        def wrapper(*args, **kwargs):
            now = time.time()
            if key in _cache and now - _cache[key]['time'] < ttl:
                return _cache[key]['data']
            result = func(*args, **kwargs)
            _cache[key] = {'data': result, 'time': now}
            return result
        return wrapper
    return decorator

def get_db_connection():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def get_visitor_count():
    try:
        with open(VISITOR_PATH, 'r') as f:
            data = json.load(f)
            return data.get('count', 0)
    except:
        return 0

def increment_visitor():
    count = get_visitor_count() + 1
    with open(VISITOR_PATH, 'w') as f:
        json.dump({'count': count}, f)
    return count

def get_last_update():
    try:
        mt = os.path.getmtime(DB_PATH)
        t = time.localtime(mt)
        days = ['Senin', 'Selasa', 'Rabu', 'Kamis', 'Jumat', 'Sabtu', 'Minggu']
        day_name = days[t.tm_wday]
        return f"{day_name}, {time.strftime('%Y.%m.%d jam %H.%M', t)}"
    except:
        return '--/--/--||--:--'

def parse_budget(budget_str):
    """Parse 'Rp 1,000,000' string to float"""
    if not budget_str:
        return 0
    return float(budget_str.replace('Rp ', '').replace(',', ''))


def get_year():
    try:
        return int(request.args.get('tahun', '2026'))
    except:
        return 2026

class FilterQuery:
    """Build and manage filter SQL queries"""
    def __init__(self, search_query='', m_filter='', type_filter='', cara_filter='', quick_search=''):
        self.search_query = search_query
        self.m_filter = m_filter
        self.type_filter = type_filter
        self.cara_filter = cara_filter
        self.quick_search = quick_search

    def where_clause(self, alias='p'):
        """Build WHERE clause and return (clause_string, params_list)"""
        parts = []
        params = []
        prefix = f'{alias}.' if alias else ''

        if self.search_query:
            parts.append(f"({prefix}package_name LIKE ? OR {prefix}satker LIKE ? OR CAST({prefix}id AS TEXT) LIKE ?)")
            sq = f'%{self.search_query}%'
            params.extend([sq, sq, sq])

        # Quick search: narrows within results already filtered by search_query
        if self.quick_search:
            parts.append(f"({prefix}package_name LIKE ? OR {prefix}satker LIKE ? OR CAST({prefix}id AS TEXT) LIKE ?)")
            qs = f'%{self.quick_search}%'
            params.extend([qs, qs, qs])

        if self.m_filter:
            parts.append(f"{prefix}procurement_method = ?")
            params.append(self.m_filter)

        if self.type_filter:
            parts.append(f"{prefix}procurement_type = ?")
            params.append(self.type_filter)

        if self.cara_filter:
            parts.append(f"{prefix}\"Cara Pengadaan\" = ?")
            params.append(self.cara_filter)

        if parts:
            return ' AND '.join(parts), params
        return '', []

    @property
    def has_filters(self):
        return bool(self.search_query or self.m_filter or self.type_filter or self.cara_filter)


def get_threshold_case():
    """SQL CASE expression for PL threshold per procurement type"""
    return ("CASE WHEN procurement_type = 'Jasa Konsultansi' THEN 100000000 "
            "WHEN procurement_type = 'Pekerjaan Konstruksi' THEN 400000000 "
            "ELSE 200000000 END")


def fetch_table_data(conn, search_query, m_filter, type_filter, cara_filter='', sort_by='id', sort_dir='DESC', page=1, quick_search='', tahun=2026):
    """Fetch paginated table data + stats. Returns dict with all needed data."""
    fq = FilterQuery(search_query, m_filter, type_filter, cara_filter, quick_search)
    offset = (page - 1) * PER_PAGE

    allowed_sorts = {
        'id': 'p.id',
        'package_name': 'p.package_name',
        'satker': 'p.satker',
        'budget': BUDGET_SQL
    }
    sort_col = allowed_sorts.get(sort_by, 'p.id')
    sort_dir_sql = 'ASC' if sort_dir.upper() == 'ASC' else 'DESC'


    # Build WHERE with tahun filter
    where_clause, params = fq.where_clause()
    tw = '= ?'
    yp = [tahun]
    if where_clause:
        where_clause += f' AND p."Tahun Anggaran" {tw}'
        params += yp
    else:
        where_clause = f'p."Tahun Anggaran" {tw}'
        params = yp

    # --- COUNT ---
    count_q = f"SELECT COUNT(*) as cnt FROM procurement p WHERE {where_clause}"
    filtered_count = conn.execute(count_q, params).fetchone()['cnt'] or 0

    # --- DATA (optimized: LEFT JOIN instead of correlated subquery) ---
    data_q = f"""SELECT p.*,
        COALESCE(rt.total_real, 0) as realisasi_total
    FROM procurement p
    LEFT JOIN (
        SELECT "Kode RUP" as kr, SUM("Total Nilai (Rp)") as total_real
        FROM realisasi WHERE "Tahun Anggaran" = {tahun} GROUP BY "Kode RUP"
    ) rt ON rt.kr = CAST(p.id AS TEXT)
    WHERE {where_clause}
    ORDER BY {sort_col} {sort_dir_sql} LIMIT {PER_PAGE} OFFSET {offset}"""

    rows = conn.execute(data_q, params).fetchall()

    packages = []
    page_total_budget = 0
    for row in rows:
        pkg = dict(row)
        budget_num = parse_budget(pkg.get('budget'))
        pkg_real = pkg.get('realisasi_total', 0) or 0
        pkg['has_realisasi'] = pkg_real > 0
        pkg['realisasi_persen'] = round((pkg_real / budget_num) * 100, 1) if budget_num > 0 and pkg_real > 0 else 0
        # Determine PL over-limit for this package
        if pkg['procurement_method'] == 'Pengadaan Langsung':
            if pkg['procurement_type'] == 'Pekerjaan Konstruksi':
                pkg['_is_pl_over'] = budget_num > 400000000
            elif pkg['procurement_type'] == 'Jasa Konsultansi':
                pkg['_is_pl_over'] = budget_num > 100000000
            else:
                pkg['_is_pl_over'] = budget_num > 200000000
        else:
            pkg['_is_pl_over'] = False
        pkg['_row_num'] = offset + rows.index(row) + 1
        page_total_budget += budget_num
        packages.append(pkg)

    total_pages = max(1, (filtered_count + PER_PAGE - 1) // PER_PAGE)

    # --- FILTERED BUDGET ---
    fb_q = f"SELECT SUM({BUDGET_SQL}) FROM procurement p WHERE {where_clause}"
    filtered_total_budget = conn.execute(fb_q, params).fetchone()[0] or 0

    # --- FILTERED BREAKDOWN ---
    f_detail_q = f"""
        SELECT
            CASE WHEN procurement_method IN ('E-Purchasing', 'Pengadaan Langsung', 'Dikecualikan', 'Seleksi')
                 THEN procurement_method ELSE 'Lainnya' END as method_group,
            COUNT(*) as total, SUM({BUDGET_SQL}) as total_budget
        FROM procurement p
        WHERE {where_clause}
        GROUP BY method_group"""

    f_detail = conn.execute(f_detail_q, params).fetchall()
    fd = {'ep': 0, 'bd_ep': 0, 'pl': 0, 'bd_pl': 0, 'dk': 0, 'bd_dk': 0, 'sl': 0, 'bd_sl': 0, 'ln': 0, 'bd_ln': 0}
    for row in f_detail:
        g, cnt, bgt = row['method_group'], row['total'], row['total_budget'] or 0
        if g == 'E-Purchasing': fd['ep'] = cnt; fd['bd_ep'] = bgt
        elif g == 'Pengadaan Langsung': fd['pl'] = cnt; fd['bd_pl'] = bgt
        elif g == 'Dikecualikan': fd['dk'] = cnt; fd['bd_dk'] = bgt
        elif g == 'Seleksi': fd['sl'] = cnt; fd['bd_sl'] = bgt
        else: fd['ln'] = cnt; fd['bd_ln'] = bgt

    return {
        'packages': packages,
        'filtered_count': filtered_count,
        'total_pages': total_pages,
        'current_page': page,
        'filtered_total_budget': filtered_total_budget,
        'page_total_budget': page_total_budget,
        **fd
    }


@app.route('/')
def index():
    tahun = get_year()
    total_visits = increment_visitor()
    last_update = get_last_update()

    search_query = request.args.get('search', '')
    m_filter = request.args.get('method', '')
    type_filter = request.args.get('type', '')
    cara_filter = request.args.get('cara', '')
    sort_by = request.args.get('sort_by', 'id')
    sort_dir = request.args.get('sort_dir', 'DESC')
    page = int(request.args.get('page', 1))

    conn = get_db_connection()

    # Global stats (always the same, independent of filter)
    stats = conn.execute(
        f"SELECT COUNT(*) as total, SUM({BUDGET_SQL}) as total_budget FROM procurement"
    ).fetchone()

    detail_grouped = conn.execute(f"""
        SELECT
            CASE WHEN procurement_method IN ('E-Purchasing', 'Pengadaan Langsung', 'Dikecualikan', 'Seleksi')
                 THEN procurement_method ELSE 'Lainnya' END as method_group,
            COUNT(*) as total, SUM({BUDGET_SQL}) as total_budget
        FROM procurement GROUP BY method_group
    """).fetchall()

    total_epurchasing = 0; budget_epurchasing = 0
    total_pl = 0; budget_pl = 0
    total_dikecualikan = 0; budget_dikecualikan = 0
    total_seleksi = 0; budget_seleksi = 0
    total_lainnya = 0; budget_lainnya = 0
    for row in detail_grouped:
        g, cnt, bgt = row['method_group'], row['total'], row['total_budget'] or 0
        if g == 'E-Purchasing': total_epurchasing = cnt; budget_epurchasing = bgt
        elif g == 'Pengadaan Langsung': total_pl = cnt; budget_pl = bgt
        elif g == 'Dikecualikan': total_dikecualikan = cnt; budget_dikecualikan = bgt
        elif g == 'Seleksi': total_seleksi = cnt; budget_seleksi = bgt
        else: total_lainnya = cnt; budget_lainnya = bgt

    # Realisasi global
    real_total = conn.execute('SELECT SUM("Total Nilai (Rp)") FROM realisasi WHERE "Tahun Anggaran" = ?', (tahun,)).fetchone()[0] or 0
    count_realized = conn.execute('SELECT COUNT(*) FROM realisasi WHERE "Tahun Anggaran" = ?', (tahun,)).fetchone()[0] or 0

    real_grouped = conn.execute(f"""
        SELECT CASE WHEN "Metode Pengadaan" IN ('E-Purchasing', 'Pengadaan Langsung', 'Dikecualikan', 'Seleksi') THEN "Metode Pengadaan" ELSE 'Lainnya' END as method_group, SUM("Total Nilai (Rp)") as total_budget
        FROM realisasi WHERE "Tahun Anggaran" = {tahun} GROUP BY method_group
    """).fetchall()
    real_epurchasing = 0; real_pl = 0; real_dikecualikan = 0; real_seleksi = 0; real_lainnya = 0
    for row in real_grouped:
        g, bgt = row['method_group'], row['total_budget'] or 0
        if g == 'E-Purchasing': real_epurchasing = bgt
        elif g == 'Pengadaan Langsung': real_pl = bgt
        elif g == 'Dikecualikan': real_dikecualikan = bgt
        elif g == 'Seleksi': real_seleksi = bgt
        else: real_lainnya = bgt

    # Table data (filtered/paginated)
    td = fetch_table_data(conn, search_query, m_filter, type_filter, cara_filter, sort_by, sort_dir, page, tahun=tahun)

    # Methods & types for filter dropdowns
    methods = conn.execute(
        "SELECT DISTINCT procurement_method FROM procurement WHERE procurement_method IS NOT NULL AND \"Tahun Anggaran\" = ?", (tahun,)
    ).fetchall()
    types = conn.execute(
        "SELECT DISTINCT procurement_type FROM procurement WHERE procurement_type IS NOT NULL AND \"Tahun Anggaran\" = ?", (tahun,)
    ).fetchall()
    cara_list = conn.execute(
        '''SELECT DISTINCT "Cara Pengadaan" FROM procurement WHERE "Cara Pengadaan" IS NOT NULL AND "Cara Pengadaan" != '' '''
    ).fetchall()

    # Suggestions
    suggestions = {
        "package_names": conn.execute("SELECT DISTINCT package_name FROM procurement WHERE \"Tahun Anggaran\" = ?", (tahun,)).fetchall(),
        "satkers": conn.execute("SELECT DISTINCT satker FROM procurement WHERE \"Tahun Anggaran\" = ?", (tahun,)).fetchall(),
        "kode_rup": conn.execute("SELECT DISTINCT id FROM procurement WHERE \"Tahun Anggaran\" = ? ORDER BY id", (tahun,)).fetchall()
    }

    # Anomalies
    pl_anomaly_count = conn.execute(f"""
        SELECT COUNT(*) FROM procurement
        WHERE procurement_method = 'Pengadaan Langsung'
        AND "Tahun Anggaran" = {tahun}
        AND {BUDGET_SQL} > {get_threshold_case()}
    """).fetchone()[0] or 0

    pl_anomalies = conn.execute(f"""
        SELECT id, package_name, satker, budget, procurement_type, work_description,
               'Pengadaan Langsung' as procurement_method,
               {get_threshold_case()} as threshold
        FROM procurement
        WHERE procurement_method = 'Pengadaan Langsung'
        AND "Tahun Anggaran" = {tahun}
        AND {BUDGET_SQL} > {get_threshold_case()}
        ORDER BY {BUDGET_SQL} DESC
    """).fetchall()

    pagu_anomaly_count = conn.execute(f"""
        SELECT COUNT(*) FROM procurement p
        WHERE p."Tahun Anggaran" = {tahun} AND {BUDGET_SQL} > 0
        AND (SELECT SUM(r2."Total Nilai (Rp)") FROM realisasi r2 WHERE r2."Kode RUP" = CAST(p.id AS TEXT) AND r2."Tahun Anggaran" = {tahun}) > {BUDGET_SQL}
    """).fetchone()[0] or 0

    pagu_anomalies = conn.execute(f"""
        SELECT p.id, p.package_name, p.satker, p.budget, p.procurement_method, p.procurement_type, p.work_description,
               (SELECT SUM(r2."Total Nilai (Rp)") FROM realisasi r2 WHERE r2."Kode RUP" = CAST(p.id AS TEXT) AND r2."Tahun Anggaran" = {tahun}) as realisasi_total
        FROM procurement p
        WHERE p."Tahun Anggaran" = {tahun} AND {BUDGET_SQL} > 0
        AND (SELECT SUM(r2."Total Nilai (Rp)") FROM realisasi r2 WHERE r2."Kode RUP" = CAST(p.id AS TEXT) AND r2."Tahun Anggaran" = {tahun}) > {BUDGET_SQL}
        ORDER BY realisasi_total DESC
    """).fetchall()

    # Splash info: RUP tanpa realisasi
    splash1 = conn.execute(f"""
        SELECT COUNT(*) as cnt, COALESCE(SUM({BUDGET_SQL}), 0) as total
        FROM procurement p
        WHERE p."Tahun Anggaran" = {tahun}
        AND (SELECT COUNT(*) FROM realisasi r WHERE r."Kode RUP" = CAST(p.id AS TEXT) AND r."Tahun Anggaran" = {tahun}) = 0
    """).fetchone()
    splash_no_realisasi_cnt = splash1['cnt']
    splash_no_realisasi_total = splash1['total']

    # Splash info: realisasi tanpa RUP awal
    splash2 = conn.execute(f"""
        SELECT COUNT(*) as cnt, COALESCE(SUM(r."Total Nilai (Rp)"), 0) as total
        FROM realisasi r
        WHERE r."Tahun Anggaran" = {tahun}
        AND r."Kode RUP" NOT IN (SELECT id FROM procurement WHERE "Tahun Anggaran" = {tahun})
    """).fetchone()
    splash_orphan_real_cnt = splash2['cnt']
    splash_orphan_real_total = splash2['total']

    conn.close()

    return render_template('index.html',
        # Global
        total_count=stats['total'],
        total_budget=stats['total_budget'],
        total_epurchasing=total_epurchasing, budget_epurchasing=budget_epurchasing,
        total_pl=total_pl, budget_pl=budget_pl,
        total_dikecualikan=total_dikecualikan, budget_dikecualikan=budget_dikecualikan,
        total_seleksi=total_seleksi, budget_seleksi=budget_seleksi,
        total_lainnya=total_lainnya, budget_lainnya=budget_lainnya,
        real_total=real_total, count_realized=count_realized,
        real_epurchasing=real_epurchasing, real_pl=real_pl,
        real_dikecualikan=real_dikecualikan, real_seleksi=real_seleksi,
        real_lainnya=real_lainnya,
        # Filtered
        packages=td['packages'],
        filtered_count=td['filtered_count'],
        total_pages=td['total_pages'],
        filtered_total_budget=td['filtered_total_budget'],
        page_total_budget=td['page_total_budget'],
        f_total_epurchasing=td['ep'], f_budget_epurchasing=td['bd_ep'],
        f_total_pl=td['pl'], f_budget_pl=td['bd_pl'],
        f_total_dikecualikan=td['dk'], f_budget_dikecualikan=td['bd_dk'],
        f_total_seleksi=td['sl'], f_budget_seleksi=td['bd_sl'],
        f_total_lainnya=td['ln'], f_budget_lainnya=td['bd_ln'],
        # Filters
        search_query=search_query, m_filter=m_filter, type_filter=type_filter, cara_filter=cara_filter,
        sort_by=sort_by, sort_dir=sort_dir, page=page, per_page=PER_PAGE,
        # UI data
        methods=methods, types=types, cara_list=cara_list, suggestions=suggestions,
        total_visits=total_visits, last_update=last_update,
        # Anomalies
        pl_anomalies=[dict(r) for r in pl_anomalies],
        pl_anomaly_count=pl_anomaly_count,
        pagu_anomalies=[dict(r) for r in pagu_anomalies],
        pagu_anomaly_count=pagu_anomaly_count,
        # Splash info
        splash_no_realisasi_cnt=splash_no_realisasi_cnt,
        splash_no_realisasi_total=splash_no_realisasi_total,
        splash_orphan_real_cnt=splash_orphan_real_cnt,
        splash_orphan_real_total=splash_orphan_real_total
    )


@app.route('/api/table')
@app.route('/api/table')
def api_table():
    """API endpoint returning JSON for AJAX table refresh."""
    search_query = request.args.get('search', '')
    m_filter = request.args.get('method', '')
    type_filter = request.args.get('type', '')
    cara_filter = request.args.get('cara', '')
    quick_search = request.args.get('quick_search', '')
    sort_by = request.args.get('sort_by', 'id')
    sort_dir = request.args.get('sort_dir', 'DESC')
    page = int(request.args.get('page', 1))
    tahun = get_year()

    conn = get_db_connection()
    td = fetch_table_data(conn, search_query, m_filter, type_filter, cara_filter, sort_by, sort_dir, page, quick_search, tahun=tahun)
    conn.close()

    return jsonify({
        'success': True,
        'packages': td['packages'],
        'filtered_count': td['filtered_count'],
        'total_pages': td['total_pages'],
        'current_page': td['current_page'],
        'filtered_total_budget': td['filtered_total_budget'],
        'page_total_budget': td['page_total_budget'],
        'f_total_epurchasing': td['ep'],
        'f_budget_epurchasing': td['bd_ep'],
        'f_total_pl': td['pl'],
        'f_budget_pl': td['bd_pl'],
        'f_total_dikecualikan': td['dk'],
        'f_budget_dikecualikan': td['bd_dk'],
        'f_total_seleksi': td['sl'],
        'f_budget_seleksi': td['bd_sl'],
        'f_total_lainnya': td['ln'],
        'f_budget_lainnya': td['bd_ln'],
    })


@app.route('/export')
def export():
    search_query = request.args.get('search', '')
    m_filter = request.args.get('method', '')
    type_filter = request.args.get('type', '')
    cara_filter = request.args.get('cara', '')
    file_format = request.args.get('format', 'xlsx')
    tahun = get_year()

    conn = get_db_connection()
    fq = FilterQuery(search_query, m_filter, type_filter, cara_filter)
    query = "SELECT * FROM procurement WHERE 1=1"
    where_clause, params = fq.where_clause(alias='')
    if where_clause:
        query += f" AND {where_clause}"
    params.append(tahun)
    query += ' AND "Tahun Anggaran" = ?'
    df = pd.read_sql_query(query, conn, params=params)
    conn.close()

    output = BytesIO()
    if file_format == 'xlsx':
        cols_to_export = ['id', 'package_name', 'satker', 'procurement_method', 'procurement_type', 'budget', 'work_description']
        df_export = df[cols_to_export].copy()
        df_export.columns = ['Kode RUP', 'Nama Paket', 'Satker', 'Metode', 'Tipe', 'Anggaran', 'Deskripsi']

        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
            df_export.to_excel(writer, index=False, sheet_name='Data Pengadaan')
            worksheet = writer.sheets['Data Pengadaan']
            worksheet.set_column('A:A', 12)
            worksheet.set_column('B:B', 40)
            worksheet.set_column('C:C', 35)
            worksheet.set_column('D:E', 15)
            worksheet.set_column('F:F', 20)
            worksheet.set_column('G:G', 50)

        mime = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        filename = "sirup_bulukumba_export.xlsx"

    elif file_format == 'pdf':
        doc = SimpleDocTemplate(output, pagesize=landscape(A4), rightMargin=20, leftMargin=20, topMargin=20, bottomMargin=20)
        elements = []
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle('TitleStyle', parent=styles['Heading1'], alignment=1, spaceAfter=20)
        elements.append(Paragraph("Laporan Data Pengadaan SIRUP Bulukumba", title_style))
        data = [['Kode RUP', 'Nama Paket', 'Satker', 'Metode', 'Anggaran']]
        df_pdf = df.head(500)
        for _, row in df_pdf.iterrows():
            pkg = Paragraph(str(row['package_name'])[:100] + ('...' if len(str(row['package_name']))>100 else ''), styles['Normal'])
            satker = Paragraph(str(row['satker'])[:80] + ('...' if len(str(row['satker']))>80 else ''), styles['Normal'])
            budget = str(row['budget'])
            data.append([str(row['id']), pkg, satker, str(row['procurement_method']), budget])
        table = Table(data, colWidths=[70, 250, 200, 100, 100])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4f46e5')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.lightgrey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')])
        ]))
        elements.append(table)
        if len(df) > 500:
            elements.append(Spacer(1, 10))
            elements.append(Paragraph(
                "* Catatan: Laporan PDF ini dibatasi 500 data pertama dari total {} data. "
                "Silakan gunakan filter atau unduh Excel untuk melihat selengkapnya.".format(len(df)),
                styles['Italic']))
        doc.build(elements)
        mime = 'application/pdf'
        filename = "sirup_bulukumba_export.pdf"
    else:
        df.to_csv(output, index=False, encoding='utf-8-sig')
        mime = 'text/csv'
        filename = "sirup_bulukumba_export.csv"

    output.seek(0)
    return send_file(output, mimetype=mime, as_attachment=True, download_name=filename)


@app.route('/admin/upload', methods=['POST'])
def admin_upload():
    password = request.form.get('password')
    if password != 'Callunk13':
        return jsonify({'success': False, 'message': 'Password salah!'}), 403

    rup_files = request.files.getlist('rup_file')
    realisasi_files = request.files.getlist('realisasi_file')

    if len(rup_files) == 0 and len(realisasi_files) == 0:
        return jsonify({'success': False, 'message': 'Minimal satu file harus diupload!'}), 400
    if all(f.filename == '' for f in rup_files + realisasi_files):
        return jsonify({'success': False, 'message': 'Minimal satu file harus diupload!'}), 400

    def _detect_year(df, default=2026):
        if 'Tahun Anggaran' in df.columns:
            s = df['Tahun Anggaran'].dropna()
            if not s.empty:
                return int(s.mode().iloc[0])
        if 'Tahun' in df.columns:
            s = df['Tahun'].dropna()
            if not s.empty:
                return int(s.mode().iloc[0])
        return default

    def _process_rup(df, conn):
        tahun = _detect_year(df)
        conn.execute('DELETE FROM procurement WHERE "Tahun Anggaran" = ?', (tahun,))
        col_map = {
            'Kode RUP': 'id', 'Nama Satuan Kerja': 'satker', 'Nama Paket': 'package_name',
            'Metode Pengadaan': 'procurement_method', 'Jenis Pengadaan': 'procurement_type',
            'Total Nilai (Rp)': 'budget', 'Sumber Dana': 'funding_source', 'Produk Dalam Negeri': 'is_umkm'
        }
        if 'Kode RUP' in df.columns:
            before = len(df)
            df = df.drop_duplicates(subset=['Kode RUP'])
            after = len(df)
            if before > after:
                print(f'[DEDUP] Removed {before - after} duplicate Kode RUP rows')
        rename_dict = {k: v for k, v in col_map.items() if k in df.columns}
        df = df.rename(columns=rename_dict)
        if 'work_description' not in df.columns:
            df['work_description'] = ''
        if 'risk_score' not in df.columns:
            df['risk_score'] = 0
        if 'budget' in df.columns and pd.api.types.is_numeric_dtype(df['budget']):
            df['budget'] = df['budget'].apply(lambda x: f"Rp {x:,.0f}" if pd.notnull(x) else "Rp 0")
        df.to_sql('procurement', conn, if_exists='append', index=False)
        conn.execute('DELETE FROM procurement WHERE rowid NOT IN (SELECT MIN(rowid) FROM procurement GROUP BY id, "Tahun Anggaran")')
        print(f'[UPLOAD] Inserted {len(df)} procurement records for {tahun}')
        return tahun

    def _process_realisasi(df, conn):
        tahun = _detect_year(df)
        conn.execute('DELETE FROM realisasi WHERE "Tahun Anggaran" = ?', (tahun,))
        df.to_sql('realisasi', conn, if_exists='append', index=False)
        print(f'[UPLOAD] Inserted {len(df)} realisasi records for {tahun}')
        return tahun

    try:
        conn = sqlite3.connect(DB_PATH)
        years_processed = []

        for f in rup_files:
            if f and f.filename:
                df = pd.read_csv(f, low_memory=False)
                y = _process_rup(df, conn)
                years_processed.append(f'RUP {y}')

        for f in realisasi_files:
            if f and f.filename:
                df = pd.read_csv(f, low_memory=False)
                y = _process_realisasi(df, conn)
                years_processed.append(f'Realisasi {y}')

        conn.commit()
        conn.close()
        msg = f'Database berhasil diupdate! ({len(years_processed)} file: {", ".join(years_processed)})'
        return jsonify({'success': True, 'message': msg})
    except Exception as e:
        conn.close()
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/package_details')
def package_details():
    package_name = request.args.get('package_name', '')
    satker = request.args.get('satker', '')
    p_id = request.args.get('id', '')

    conn = get_db_connection()
    try:
        package_id = str(p_id) if p_id else ""
        query = 'SELECT * FROM realisasi WHERE "Kode RUP" = ?'
        realisasi = conn.execute(query, (package_id,)).fetchall()

        result = [dict(row) for row in realisasi]
        return jsonify({'success': True, 'data': result})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})
    finally:
        conn.close()


@cached('global_stats')
def _get_global_stats():
    conn = get_db_connection()
    stats = conn.execute(
        f"SELECT COUNT(*) as total, SUM({BUDGET_SQL}) as total_budget FROM procurement"
    ).fetchone()

    detail_grouped = conn.execute(f"""
        SELECT
            CASE WHEN procurement_method IN ('E-Purchasing', 'Pengadaan Langsung', 'Dikecualikan', 'Seleksi')
                 THEN procurement_method ELSE 'Lainnya' END as method_group,
            COUNT(*) as total, SUM({BUDGET_SQL}) as total_budget
        FROM procurement GROUP BY method_group
    """).fetchall()

    real_total = conn.execute(
        'SELECT SUM("Total Nilai (Rp)") as total FROM realisasi'
    ).fetchone()['total'] or 0
    count_realized = conn.execute(
        'SELECT COUNT(DISTINCT "Kode RUP") as cnt FROM realisasi'
    ).fetchone()['cnt'] or 0

    def _find(dg, name):
        for r in dg:
            if r['method_group'] == name:
                return r['total'], r['total_budget'] or 0
        return 0, 0

    return {
        'total_count': stats['total'],
        'total_budget': stats['total_budget'] or 0,
        'ep': dict(zip(['total','budget'], _find(detail_grouped, 'E-Purchasing'))),
        'pl': dict(zip(['total','budget'], _find(detail_grouped, 'Pengadaan Langsung'))),
        'dk': dict(zip(['total','budget'], _find(detail_grouped, 'Dikecualikan'))),
        'sl': dict(zip(['total','budget'], _find(detail_grouped, 'Seleksi'))),
        'ln': dict(zip(['total','budget'], _find(detail_grouped, 'Lainnya'))),
        'real_total': real_total,
        'count_realized': count_realized,
        'last_update': get_last_update(),
    }


@app.route('/api/stats')
def api_stats():
    return jsonify(_get_global_stats())


@app.route('/api/anomalies/pagu')
def api_anomalies_pagu():
    """Return pagu anomalies. Optimized with LEFT JOIN."""
    tahun = get_year()
    conn = get_db_connection()
    B = """CAST(REPLACE(REPLACE(p.budget, 'Rp ', ''), ',', '') AS REAL)"""
    count_q = f"""SELECT COUNT(*) FROM procurement p
        LEFT JOIN (
            SELECT "Kode RUP" as kr, SUM("Total Nilai (Rp)") as total_real
            FROM realisasi WHERE "Tahun Anggaran" = {tahun} GROUP BY "Kode RUP"
        ) r ON r.kr = CAST(p.id AS TEXT)
        WHERE p."Tahun Anggaran" = {tahun} AND {B} > 0
        AND COALESCE(r.total_real, 0) > {B}"""
    count = conn.execute(count_q).fetchone()[0] or 0
    data_q = f"""SELECT p.id, p.package_name, p.satker, p.budget,
            p.procurement_method, p.procurement_type, p.work_description,
            p."Cara Pengadaan",
            COALESCE(r.total_real, 0) as realisasi_total
        FROM procurement p
        LEFT JOIN (
            SELECT "Kode RUP" as kr, SUM("Total Nilai (Rp)") as total_real
            FROM realisasi WHERE "Tahun Anggaran" = {tahun} GROUP BY "Kode RUP"
        ) r ON r.kr = CAST(p.id AS TEXT)
        WHERE p."Tahun Anggaran" = {tahun} AND {B} > 0
        AND COALESCE(r.total_real, 0) > {B}
        ORDER BY realisasi_total DESC"""
    rows = [dict(r) for r in conn.execute(data_q).fetchall()]
    conn.close()
    return jsonify({
        'success': True, 'count': count, 'data': rows,
        'mode': 'strict', 'match_label': 'Kode RUP'
    })

@app.route('/api/anomalies/orphan')
def api_anomalies_orphan():
    """Return orphan realisasi (no parent RUP in procurement table)"""
    tahun = get_year()
    conn = get_db_connection()
    count = conn.execute(
        'SELECT COUNT(*) FROM realisasi WHERE "Tahun Anggaran" = ? AND "Kode RUP" NOT IN (SELECT id FROM procurement WHERE "Tahun Anggaran" = ?)',
        (tahun, tahun)
    ).fetchone()[0] or 0
    rows = conn.execute('''
        SELECT "Kode RUP" as kode_rup, "Nama Paket" as nama_paket,
               "Nama Satuan Kerja" as satker, "Nama Penyedia" as penyedia,
               "Total Nilai (Rp)" as total_nilai, "Metode Pengadaan" as metode,
               "Jenis Pengadaan" as jenis, "Tahun Anggaran" as tahun,
               "Status Paket" as status_paket
        FROM realisasi
        WHERE "Tahun Anggaran" = ? AND "Kode RUP" NOT IN (SELECT id FROM procurement WHERE "Tahun Anggaran" = ?)
        ORDER BY "Nama Paket" ASC
    ''', (tahun, tahun)).fetchall()
    conn.close()
    return jsonify({
        'success': True,
        'count': count,
        'data': [dict(r) for r in rows]
    })


@app.route('/penyedia/export')
def export_penyedia():
    search_query = request.args.get('search', '')
    verif_filter = request.args.get('verif', '')
    file_format = request.args.get('format', 'xlsx')

    conn = get_db_connection()
    query = "SELECT nama, npwp, email, tanggal_daftar, jenis_usaha, alamat, telepon, terverifikasi FROM penyedia WHERE 1=1"
    params = []
    if search_query:
        query += " AND (nama LIKE ? OR npwp LIKE ? OR jenis_usaha LIKE ? OR alamat LIKE ?)"
        sq = f'%{search_query}%'
        params.extend([sq, sq, sq, sq])
    if verif_filter == 'verified':
        query += " AND terverifikasi = 1"
    elif verif_filter == 'unverified':
        query += " AND (terverifikasi IS NULL OR terverifikasi = 0)"
    elif verif_filter == 'pt':
        query += " AND jenis_usaha LIKE '%PT%'"
    elif verif_filter == 'perorangan':
        query += " AND (jenis_usaha LIKE '%Perorangan%' OR jenis_usaha LIKE '%Usaha Perorangan%')"
    query += " ORDER BY nama ASC"
    df = pd.read_sql_query(query, conn, params=params)
    conn.close()

    output = BytesIO()
    if file_format == 'xlsx':
        df_export = df.copy()
        df_export.columns = ['Nama Penyedia', 'NPWP', 'Email', 'Tanggal Daftar', 'Jenis Usaha', 'Alamat', 'Telepon', 'Terverifikasi']
        df_export['Terverifikasi'] = df_export['Terverifikasi'].apply(lambda x: '✅ Ya' if x == 1 else '⬜ Belum')
        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
            df_export.to_excel(writer, index=False, sheet_name='Penyedia')
            worksheet = writer.sheets['Penyedia']
            worksheet.set_column('A:A', 35)
            worksheet.set_column('B:B', 22)
            worksheet.set_column('C:C', 30)
            worksheet.set_column('D:D', 18)
            worksheet.set_column('E:E', 22)
            worksheet.set_column('F:F', 45)
            worksheet.set_column('G:G', 20)
            worksheet.set_column('H:H', 14)
        mime = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        filename = 'penyedia_bulukumba.xlsx'
    elif file_format == 'pdf':
        doc = SimpleDocTemplate(output, pagesize=landscape(A4), rightMargin=20, leftMargin=20, topMargin=20, bottomMargin=20)
        elements = []
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle('TitleStyle', parent=styles['Heading1'], alignment=1, spaceAfter=20, textColor=colors.HexColor('#059669'))
        elements.append(Paragraph('Daftar Penyedia Terverifikasi Kab. Bulukumba', title_style))
        data = [['No', 'Nama Penyedia', 'NPWP', 'Jenis Usaha', 'Status']]
        df_pdf = df.head(1000)
        for idx, (_, row) in enumerate(df_pdf.iterrows(), 1):
            status = '✅ Terverifikasi' if row['terverifikasi'] == 1 else '⬜ Belum'
            nama = Paragraph(str(row['nama'])[:60], styles['Normal'])
            data.append([str(idx), nama, str(row['npwp']), str(row['jenis_usaha']), status])
        table = Table(data, colWidths=[30, 280, 160, 120, 80])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#059669')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.lightgrey),
        ]))
        elements.append(table)
        if len(df) > 1000:
            elements.append(Spacer(1, 10))
            elements.append(Paragraph(f'* Menampilkan 1000 data pertama dari total {len(df)} data.', styles['Italic']))
        doc.build(elements)
        mime = 'application/pdf'
        filename = 'penyedia_bulukumba.pdf'
    else:
        df.to_csv(output, index=False, encoding='utf-8-sig')
        mime = 'text/csv'
        filename = 'penyedia_bulukumba.csv'
    output.seek(0)
    return send_file(output, mimetype=mime, as_attachment=True, download_name=filename)


@app.route('/penyedia')
def list_penyedia():
    search_query = request.args.get('search', '')
    verif_filter = request.args.get('verif', '')
    page = int(request.args.get('page', 1))
    per_page = 50
    offset = (page - 1) * per_page

    conn = get_db_connection()
    count_query = "SELECT COUNT(*) FROM penyedia WHERE 1=1"
    data_query = "SELECT * FROM penyedia WHERE 1=1"
    params = []
    if search_query:
        count_query += " AND (nama LIKE ? OR npwp LIKE ? OR jenis_usaha LIKE ? OR alamat LIKE ?)"
        data_query += " AND (nama LIKE ? OR npwp LIKE ? OR jenis_usaha LIKE ? OR alamat LIKE ?)"
        sq = f'%{search_query}%'
        params.extend([sq, sq, sq, sq])
    if verif_filter == 'verified':
        count_query += " AND terverifikasi = 1"
        data_query += " AND terverifikasi = 1"
    elif verif_filter == 'unverified':
        count_query += " AND (terverifikasi IS NULL OR terverifikasi = 0)"
        data_query += " AND (terverifikasi IS NULL OR terverifikasi = 0)"
    elif verif_filter == 'pt':
        count_query += " AND jenis_usaha LIKE '%PT%'"
        data_query += " AND jenis_usaha LIKE '%PT%'"
    elif verif_filter == 'perorangan':
        count_query += " AND (jenis_usaha LIKE '%Perorangan%' OR jenis_usaha LIKE '%Usaha Perorangan%')"
        data_query += " AND (jenis_usaha LIKE '%Perorangan%' OR jenis_usaha LIKE '%Usaha Perorangan%')"
    total_count = conn.execute(count_query, params).fetchone()[0] or 0
    data_query += " ORDER BY nama ASC LIMIT ? OFFSET ?"
    data_params = params + [per_page, offset]
    vendors = conn.execute(data_query, data_params).fetchall()
    total_vendors = conn.execute("SELECT COUNT(*) FROM penyedia").fetchone()[0] or 0
    total_pt = conn.execute("SELECT COUNT(*) FROM penyedia WHERE jenis_usaha LIKE '%PT%'").fetchone()[0] or 0
    total_perorangan = conn.execute("SELECT COUNT(*) FROM penyedia WHERE jenis_usaha LIKE '%Perorangan%'").fetchone()[0] or 0
    total_lainnya = total_vendors - total_pt - total_perorangan
    total_verified = conn.execute("SELECT COUNT(*) FROM penyedia WHERE terverifikasi = 1").fetchone()[0] or 0
    conn.close()
    return render_template('penyedia.html',
                           vendors=vendors, total_count=total_count,
                           total_vendors=total_vendors, total_pt=total_pt,
                           total_perorangan=total_perorangan, total_lainnya=total_lainnya,
                           total_verified=total_verified,
                           search_query=search_query, verif_filter=verif_filter,
                           page=page, per_page=per_page)


@app.route('/penyedia/upload', methods=['POST'])
def upload_penyedia():
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': 'Tidak ada file yang diupload'})
    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'message': 'Nama file kosong'})
    conn = get_db_connection()
    marked = 0
    errors = []
    try:
        all_npwp_normalized = []
        all_names = []
        if file.filename.endswith('.csv'):
            import io
            content = file.read().decode('utf-8')
            import csv
            first_line = content.split('\n')[0].strip().lower()
            has_headers = any(kw in first_line for kw in ['npwp', 'nama', 'penyedia', 'email', 'tanggal'])
            if has_headers:
                reader = csv.DictReader(io.StringIO(content))
                for row in reader:
                    npwp = (row.get('NPWP', '') or '').strip()
                    nama = (row.get('Nama Penyedia', row.get('Nama', '')) or '').strip()
                    if npwp or nama:
                        all_npwp_normalized.append(npwp.replace('.','').replace('-','').replace(' ',''))
                        all_names.append(nama.upper().strip())
            else:
                reader = csv.reader(io.StringIO(content))
                for row in reader:
                    if len(row) >= 2:
                        nama = (row[0] or '').strip()
                        npwp = (row[1] or '').strip()
                        all_npwp_normalized.append(npwp.replace('.','').replace('-','').replace(' ',''))
                        all_names.append(nama.upper().strip())
        elif file.filename.endswith(('.xlsx', '.xls')):
            import openpyxl
            import io
            wb = openpyxl.load_workbook(io.BytesIO(file.read()), data_only=True)
            ws = wb.active
            first_row_values = [str(ws.cell(1, c).value or '').strip().lower() for c in range(1, min(ws.max_column + 1, 10))]
            has_headers = any(kw in ','.join(first_row_values) for kw in ['npwp', 'nama', 'penyedia', 'email', 'tanggal', 'usaha'])
            if has_headers:
                headers = first_row_values
                npwp_col = next((c for c, h in enumerate(headers) if 'npwp' in h), None)
                nama_col = next((c for c, h in enumerate(headers) if 'nama' in h or 'penyedia' in h), None)
                start_row = 2
            else:
                npwp_col = 1
                nama_col = 0
                start_row = 1
            for row in range(start_row, ws.max_row + 1):
                npwp = str(ws.cell(row, (npwp_col or 0) + 1).value or '').strip() if npwp_col is not None else ''
                nama = str(ws.cell(row, (nama_col or 0) + 1).value or '').strip() if nama_col is not None else ''
                if npwp or nama:
                    all_npwp_normalized.append(npwp.replace('.','').replace('-','').replace(' ',''))
                    all_names.append(nama.upper().strip())
        else:
            return jsonify({'success': False, 'message': 'Format file tidak didukung. Gunakan .csv, .xlsx, atau .xls'})
        db_npwp_set = {}
        db_name_set = {}
        all_db = conn.execute('SELECT id, nama, npwp, terverifikasi FROM penyedia').fetchall()
        for v in all_db:
            nn = (v['npwp'] or '').replace('.','').replace('-','').replace(' ','').strip()
            nm = (v['nama'] or '').upper().strip()
            if nn: db_npwp_set[nn] = v['id']
            if nm: db_name_set[nm] = v['id']
        marked = 0
        matched_ids = set()
        unmatched = []
        for i in range(len(all_npwp_normalized)):
            file_npwp = all_npwp_normalized[i]
            file_nama = all_names[i]
            if not file_nama:
                continue
            matched_id = None
            if file_npwp and file_npwp in db_npwp_set:
                matched_id = db_npwp_set[file_npwp]
            elif file_nama in db_name_set:
                matched_id = db_name_set[file_nama]
            if matched_id:
                if matched_id not in matched_ids:
                    conn.execute('UPDATE penyedia SET terverifikasi = 1 WHERE id = ?', (matched_id,))
                    matched_ids.add(matched_id)
                    marked += 1
            else:
                unmatched.append(file_nama.title())
        conn.commit()
    except Exception as e:
        conn.rollback()
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})
    finally:
        conn.close()
    conn2 = get_db_connection()
    total_verified = conn2.execute('SELECT COUNT(*) FROM penyedia WHERE terverifikasi = 1').fetchone()[0]
    conn2.close()
    msg = f'✅ {marked} penyedia ditandai Terverifikasi!'
    if unmatched:
        msg += f'\n⚠️ {len(unmatched)} penyedia tidak ditemukan di database (periksa ejaan nama): {unmatched[:5]}{".." if len(unmatched) > 5 else ""}'
    return jsonify({'success': True, 'message': msg, 'total_verified': total_verified})


@app.route('/api/penyedia')
def api_penyedia():
    search_query = request.args.get('search', '')
    verif_filter = request.args.get('verif', '')
    page = int(request.args.get('page', 1))
    per_page = 50
    offset = (page - 1) * per_page

    conn = get_db_connection()
    count_query = "SELECT COUNT(*) FROM penyedia WHERE 1=1"
    data_query = "SELECT * FROM penyedia WHERE 1=1"
    params = []
    if search_query:
        count_query += " AND (nama LIKE ? OR npwp LIKE ? OR jenis_usaha LIKE ? OR alamat LIKE ?)"
        data_query += " AND (nama LIKE ? OR npwp LIKE ? OR jenis_usaha LIKE ? OR alamat LIKE ?)"
        sq = f'%{search_query}%'
        params.extend([sq, sq, sq, sq])
    if verif_filter == 'verified':
        count_query += " AND terverifikasi = 1"
        data_query += " AND terverifikasi = 1"
    elif verif_filter == 'unverified':
        count_query += " AND (terverifikasi IS NULL OR terverifikasi = 0)"
        data_query += " AND (terverifikasi IS NULL OR terverifikasi = 0)"
    elif verif_filter == 'pt':
        count_query += " AND jenis_usaha LIKE '%PT%'"
        data_query += " AND jenis_usaha LIKE '%PT%'"
    elif verif_filter == 'perorangan':
        count_query += " AND (jenis_usaha LIKE '%Perorangan%' OR jenis_usaha LIKE '%Usaha Perorangan%')"
        data_query += " AND (jenis_usaha LIKE '%Perorangan%' OR jenis_usaha LIKE '%Usaha Perorangan%')"
    total_count = conn.execute(count_query, params).fetchone()[0] or 0
    total_pages = max(1, (total_count + per_page - 1) // per_page)
    data_query += " ORDER BY nama ASC LIMIT ? OFFSET ?"
    data_params = params + [per_page, offset]
    rows = conn.execute(data_query, data_params).fetchall()
    vendors = [dict(r) for r in rows]
    total_vendors = conn.execute("SELECT COUNT(*) FROM penyedia").fetchone()[0] or 0
    total_pt = conn.execute("SELECT COUNT(*) FROM penyedia WHERE jenis_usaha LIKE '%PT%'").fetchone()[0] or 0
    total_perorangan = conn.execute("SELECT COUNT(*) FROM penyedia WHERE jenis_usaha LIKE '%Perorangan%'").fetchone()[0] or 0
    total_verified = conn.execute("SELECT COUNT(*) FROM penyedia WHERE terverifikasi = 1").fetchone()[0] or 0
    conn.close()
    return jsonify({
        'vendors': vendors,
        'total_count': total_count,
        'total_pages': total_pages,
        'current_page': page,
        'per_page': per_page,
        'total_vendors': total_vendors,
        'total_pt': total_pt,
        'total_perorangan': total_perorangan,
        'total_verified': total_verified,
        'search_query': search_query,
        'verif_filter': verif_filter,
    })


@app.route('/api/penyedia_detail')
def api_penyedia_detail():
    nama = request.args.get('nama', '').strip()
    if not nama:
        return jsonify({'success': False, 'message': 'Nama penyedia tidak diberikan'})
    conn = get_db_connection()
    try:
        # 1) Exact match
        row = conn.execute(
            'SELECT * FROM penyedia WHERE nama = ? LIMIT 1',
            (nama,)
        ).fetchone()
        if row:
            return jsonify({'success': True, 'data': dict(row)})

        # 2) LIKE partial match
        row = conn.execute(
            'SELECT * FROM penyedia WHERE nama LIKE ? LIMIT 1',
            (f'%{nama}%',)
        ).fetchone()
        if row:
            return jsonify({'success': True, 'data': dict(row)})

        # 3) Remove all spaces from both sides and compare (handles "GANES HA" vs "GANESHA")
        nama_nospace = nama.replace(' ', '')
        rows = conn.execute(
            "SELECT * FROM penyedia WHERE REPLACE(nama, ' ', '') = ? ORDER BY LENGTH(nama) ASC LIMIT 5",
            (nama_nospace,)
        ).fetchall()
        if rows:
            # Pick the one with shortest name (closest match to input)
            return jsonify({'success': True, 'data': dict(rows[0])})

        # 4) LIKE without spaces (partial) — rank by how close the name is
        rows = conn.execute(
            "SELECT * FROM penyedia WHERE REPLACE(nama, ' ', '') LIKE ? ORDER BY LENGTH(nama) ASC LIMIT 10",
            (f'%{nama_nospace}%',)
        ).fetchall()
        if rows:
            # Pick shortest name among results
            return jsonify({'success': True, 'data': dict(rows[0])})

        return jsonify({'success': False, 'message': 'Penyedia tidak ditemukan'})
    finally:
        conn.close()


@app.route('/penyedia/clear_verification', methods=['POST'])
def clear_verification():
    conn = get_db_connection()
    conn.execute('UPDATE penyedia SET terverifikasi = 0')
    conn.commit()
    conn.close()
    return jsonify({'success': True, 'message': 'Semua status verifikasi telah direset'})


@app.route('/api/konsultan_check')
def api_konsultan_check():
    tahun = get_year()
    """
    Analisa Fee Konsultan: Cocokkan paket Jasa Konsultansi dengan Pekerjaan Konstruksi.
    Menggunakan multi-level matching dengan confidence score.

    Batasan biaya (Permen PUPR):
    - Perencanaan: max ~3.5% dari pagu konstruksi
    - Pengawasan: max ~2.5% dari pagu konstruksi
    - Total konsultan (perencanaan + pengawasan): max ~5%
    """
    import re

    # --- Prefix cleaners ---
    _PREFIX_CLEAN = re.compile(
        r'^(Belanja\s+Modal\s+(?:Gedung\s+dan\s+Bangunan\s+BLUD|Bangunan\s+(?:Gedung\s+(?:Kantor|Tempat\s+Kerja\s+Lainnya)|Kesehatan)|Air\s+Irigasi\s+Lainnya)\s*[-–]\s*(?:Belanja\s+Modal\s+(?:Bangunan\s+)?(?:Gedung\s+(?:Kantor|Tempat\s+Kerja\s+Lainnya)|Kesehatan)\s*[-–]\s*)?)'
        r'|^(Konsultansi\s+(?:Perencana|Pengawasan)\s+)'
        r'|^(Perencanaan\s+)'
        r'|^(Pengawasan\s+)'
        r'|^(Jasa\s+Konsultan(?:\s+(?:Perencanaan|Pengawasan))?\s+)'
        r'|^(Biaya\s+(?:Perencanaan|Pengawasan)\s+)'
        r'|^(Desain\s+Perencanaan\s+)'
        r'|^(Konsultan\s+Perencana\s+)'
        r'', re.IGNORECASE)

    _PREFIX_CONSTR = re.compile(
        r'^(Belanja\s+Modal\s+(?:Gedung\s+dan\s+Bangunan\s+BLUD|Bangunan\s+(?:Gedung\s+(?:Kantor|Tempat\s+Kerja\s+Lainnya)|Kesehatan)|Air\s+Irigasi\s+Lainnya)\s*[-–]\s*(?:Belanja\s+Modal\s+(?:Bangunan\s+)?(?:Gedung\s+(?:Kantor|Tempat\s+Kerja\s+Lainnya)|Kesehatan)\s*[-–]\s*)?)'
        r'|^(PEMBANGUNAN\s+|Pekerjaan\s+|Pembangunan\s+|Peningkatan\s+|Perluasan\s+|Rehabilitasi\s+|Pemeliharaan/Rehabilitasi\s+|PEKERJAAN\s+|Pengadaan\s+)'
        r'|^(PEMELIHARAAN\s+)', re.IGNORECASE)

    def _clean_name(name, prefix_re):
        return prefix_re.sub('', name).strip().rstrip('; ').strip().upper()

    def _extract_project_name(cons_name):
        name_upper = cons_name.upper()
        for anchor in ['PENGERJAAN ', 'PEMBANGUNAN ', 'PEMELIHARAAN ', 'REHABILITASI ']:
            idx = name_upper.find(anchor)
            if idx >= 0:
                return name_upper[idx:].strip(), 0
        return _clean_name(cons_name, _PREFIX_CLEAN), 1

    def _match_confidence(proj_name_uc, constr_name):
        constr_uc = constr_name.upper()
        constr_clean = _clean_name(constr_name, _PREFIX_CONSTR).strip()
        # Level 1: Exact match
        if proj_name_uc == constr_clean or proj_name_uc == constr_uc:
            return 100, 'exact'
        # Level 2: Project name is prefix of construction name
        if constr_clean.startswith(proj_name_uc) or constr_uc.startswith(proj_name_uc):
            return 90, 'prefix'
        # Level 3: Contains
        if proj_name_uc in constr_uc or proj_name_uc in constr_clean:
            return 80, 'contains'
        # Level 4: Try first 30 chars as prefix
        short = proj_name_uc[:30]
        if len(short) > 15 and (constr_uc.startswith(short) or constr_clean.startswith(short)):
            return 70, 'partial_prefix'
        # Level 5: First 20 chars anywhere
        short20 = proj_name_uc[:20]
        if len(short20) > 12 and (short20 in constr_uc or short20 in constr_clean):
            return 60, 'partial'
        return 0, ''

    def _classify_tipe(nama):
        n = nama.upper()
        if 'PERENCANA' in n or n.startswith('PERENCANAAN'):
            return 'Perencanaan'
        if 'PENGAWASAN' in n:
            return 'Pengawasan'
        return 'Konsultansi'

    def _fee_flag(tipe, fee):
        t = tipe
        # Referensi: Permen PUPR No.22/PRT/M/2018 jo. Permen PUPR No.8/2023
        # Biaya Perencanaan: 2-3.5% dari nilai konstruksi
        # Biaya Pengawasan: 1.5-2.5% dari nilai konstruksi
        if t == 'Perencanaan':
            return 'danger' if fee > 4.0 else ('warning' if fee > 3.5 else 'safe')
        elif t == 'Pengawasan':
            return 'danger' if fee > 3.0 else ('warning' if fee > 2.5 else 'safe')
        else:
            return 'danger' if fee > 5.0 else ('warning' if fee > 4.0 else 'safe')

    conn = get_db_connection()
    try:
        # Fetch all consultancy & construction packages
        konsultansi = conn.execute(f"""
            SELECT id, package_name, satker, budget,
                   CAST(REPLACE(REPLACE(budget, 'Rp ', ''), ',', '') AS REAL) AS nilai,
                   procurement_method
            FROM procurement
            WHERE procurement_type = 'Jasa Konsultansi' AND "Tahun Anggaran" = {tahun}
        """).fetchall()

        konstruksi = conn.execute(f"""
            SELECT id, package_name, satker, budget,
                   CAST(REPLACE(REPLACE(budget, 'Rp ', ''), ',', '') AS REAL) AS nilai,
                   procurement_method
            FROM procurement
            WHERE procurement_type = 'Pekerjaan Konstruksi' AND "Tahun Anggaran" = {tahun}
        """).fetchall()
        conn.close()

        # Build list: only consultancy packages that look project-related
        candidates = []
        for c in konsultansi:
            d = dict(c)
            tipe = _classify_tipe(d['package_name'])
            if tipe != 'Konsultansi' or any(kw in d['package_name'].upper() for kw in ['KONSULTAN', 'DESAIN']):
                proj_name, bonus = _extract_project_name(d['package_name'])
                proj_name = proj_name.strip()
                if len(proj_name) > 5:  # Minimum meaningful length
                    candidates.append({**d, 'proj_name': proj_name, 'bonus': bonus, 'tipe': tipe})

        # Match each candidate against construction packages
        results = []
        grouped = {}

        for cand in candidates:
            best_score = 0
            best_match = None
            best_method = ''

            for k in konstruksi:
                kd = dict(k)
                if cand['satker'] != kd['satker']:
                    continue
                score, method = _match_confidence(cand['proj_name'], kd['package_name'])
                # Apply bonus for anchor-based extraction (more reliable)
                actual_score = score - cand['bonus'] * 10
                if actual_score > best_score:
                    best_score = actual_score
                    best_match = kd
                    best_method = method

            if best_match and best_score >= 50:
                constr_nilai = best_match['nilai']
                fee_persen = round(cand['nilai'] / constr_nilai * 100, 2) if constr_nilai > 0 else 0
                flag = _fee_flag(cand['tipe'], fee_persen)

                conf_label = {'100': 'Pasti', '90': 'Yakin', '80': 'Sesuai', '70': 'Mirip', '60': 'Kemungkinan'} \
                    .get(str(best_score)[:2], 'Perkiraan')

                entry = {
                    'konsultan_id': cand['id'],
                    'konsultan_nama': cand['package_name'],
                    'konsultan_budget': cand['budget'],
                    'konsultan_nilai': cand['nilai'],
                    'konsultan_metode': cand['procurement_method'],
                    'satker': cand['satker'],
                    'tipe_konsultan': cand['tipe'],
                    'fee_persen': fee_persen,
                    'flag': flag,
                    'match_score': best_score,
                    'match_method': best_method,
                    'match_label': conf_label,
                    'konstruksi_id': best_match['id'],
                    'konstruksi_nama': best_match['package_name'],
                    'konstruksi_budget': best_match['budget'],
                    'konstruksi_nilai': constr_nilai,
                    'konstruksi_metode': best_match['procurement_method'],
                }
                results.append(entry)

                # Group
                kid = best_match['id']
                if kid not in grouped:
                    grouped[kid] = {
                        'konstruksi_id': kid,
                        'konstruksi_nama': best_match['package_name'],
                        'konstruksi_nilai': constr_nilai,
                        'konstruksi_budget': best_match['budget'],
                        'konstruksi_metode': best_match['procurement_method'],
                        'satker': cand['satker'],
                        'konsultan': []
                    }
                grouped[kid]['konsultan'].append({
                    'nama': cand['package_name'],
                    'nilai': cand['nilai'],
                    'budget': cand['budget'],
                    'tipe': cand['tipe'],
                    'fee_persen': fee_persen,
                    'flag': flag,
                    'id': cand['id'],
                    'metode': cand['procurement_method'],
                    'match_label': conf_label
                })

        # Calculate totals per group
        konsolidasi = []
        for kid, g in grouped.items():
            total_fee = sum(k['nilai'] for k in g['konsultan'])
            total_persen = round(total_fee / g['konstruksi_nilai'] * 100, 2) if g['konstruksi_nilai'] > 0 else 0
            g['total_fee'] = total_fee
            g['total_persen'] = total_persen
            # Total fee perencanaan+pengawasan wajar ≤6% (Permen PUPR praktik)
            g['flag_total'] = 'danger' if total_persen > 7.0 else ('warning' if total_persen > 6.0 else 'safe')
            konsolidasi.append(g)

        konsolidasi.sort(key=lambda g: g['total_persen'], reverse=True)

        # Also find unmatched consultancies for reporting
        matched_ids = set(r['konsultan_id'] for r in results)
        unmatched = []
        for c in candidates:
            if c['id'] not in matched_ids:
                unmatched.append({
                    'id': c['id'],
                    'nama': c['package_name'],
                    'budget': c['budget'],
                    'nilai': c['nilai'],
                    'satker': c['satker'],
                    'tipe': c['tipe'],
                    'metode': c['procurement_method'],
                    'proj_name': c['proj_name']
                })

        return jsonify({
            'success': True,
            'match_count': len(results),
            'group_count': len(konsolidasi),
            'unmatched_count': len(unmatched),
            'unmatched': unmatched,
            'pairs': results,
            'grouped': konsolidasi
        })
    except Exception as e:
        conn.close()
        return jsonify({'success': False, 'message': str(e)})

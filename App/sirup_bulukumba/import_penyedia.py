import openpyxl
import sqlite3
import re
import os

DB_PATH = '/root/.openclaw/workspace/App/sirup_bulukumba/bulukumba.db'
XLSX_PATH = '/root/.openclaw/workspace/App/Penyedia_blk.xlsx'

wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
ws = wb.active

# Parse: group rows into vendor entries
vendors = []
i = 2
while i <= ws.max_row:
    nama = ws.cell(i, 1).value
    npwp = ws.cell(i, 2).value
    email = ws.cell(i, 3).value
    tgl = ws.cell(i, 4).value
    sertifikat = ws.cell(i, 5).value
    bentuk = ws.cell(i, 6).value
    
    # Skip rows that are clearly NOT vendor names (addresses, phone numbers)
    if nama and (str(nama).startswith('No. Telepon') or str(nama).startswith('JLN') or str(nama).startswith('Jl.') or str(nama).startswith('BTN') or str(nama).startswith('DUSUN') or str(nama).startswith('DESA')):
        i += 1
        continue
    
    # If it has NPWP, it's a real vendor entry
    if nama and npwp and str(npwp).strip():
        # Get address from next row
        alamat = ''
        telp = ''
        
        if i + 1 <= ws.max_row:
            next_val = ws.cell(i + 1, 1).value
            if next_val and not ws.cell(i + 1, 2).value:
                alamat = str(next_val) if next_val else ''
                # Check if row after that is phone
                if i + 2 <= ws.max_row:
                    phone_val = ws.cell(i + 2, 1).value
                    if phone_val and 'Telepon' in str(phone_val):
                        telp = str(phone_val).replace('No. Telepon : ', '').strip()
        
        jenis = str(sertifikat) if sertifikat else (str(bentuk) if bentuk else '-')
        
        vendors.append({
            'nama': str(nama).strip(),
            'npwp': str(npwp).strip(),
            'email': str(email).strip() if email else '-',
            'tanggal_daftar': str(tgl).strip() if tgl else '-',
            'jenis_usaha': jenis,
            'alamat': alamat,
            'telepon': telp
        })
    i += 1

print(f'Found {len(vendors)} vendors')

# Import to SQLite
conn = sqlite3.connect(DB_PATH)
conn.execute('''
CREATE TABLE IF NOT EXISTS penyedia (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    nama TEXT,
    npwp TEXT UNIQUE,
    email TEXT,
    tanggal_daftar TEXT,
    jenis_usaha TEXT,
    alamat TEXT,
    telepon TEXT
)
''')

# Use INSERT OR IGNORE to avoid duplicates on reruns
inserted = 0
for v in vendors:
    try:
        conn.execute('''
        INSERT OR IGNORE INTO penyedia (nama, npwp, email, tanggal_daftar, jenis_usaha, alamat, telepon)
        VALUES (?, ?, ?, ?, ?, ?, ?)
        ''', (v['nama'], v['npwp'], v['email'], v['tanggal_daftar'], v['jenis_usaha'], v['alamat'], v['telepon']))
        inserted += 1
    except Exception as e:
        pass

conn.commit()
conn.close()
print(f'Inserted {inserted} vendors into DB')

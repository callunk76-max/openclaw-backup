#!/usr/bin/env python3
"""
Extract all Pra-RKA/RKA data from Renja Kab. Bulukumba 2027
v3 - Now captures full Kode Rekening hierarchy + belanja breakdown.
"""

import sqlite3, os, re, shutil
from pathlib import Path
import openpyxl

SRC_DIR = "/tmp/renja_kab"
DB_PATH = "/root/.openclaw/workspace/App/renja_dashboard/renja.db"

BELANJA_LABELS = {
    '5.1.01': 'Belanja Pegawai',
    '5.1.02': 'Belanja Barang dan Jasa',
    '5.1.03': 'Belanja Bunga',
    '5.1.04': 'Belanja Subsidi',
    '5.1.05': 'Belanja Hibah',
    '5.1.06': 'Belanja Bantuan Sosial',
    '5.2.01': 'Belanja Modal Tanah',
    '5.2.02': 'Belanja Modal Peralatan dan Mesin',
    '5.2.03': 'Belanja Modal Gedung dan Bangunan',
    '5.2.04': 'Belanja Modal Jalan, Jaringan, dan Irigasi',
    '5.2.05': 'Belanja Modal Aset Tetap Lainnya',
}

def init_db():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.executescript("""
        DROP TABLE IF EXISTS belanja_accounts;
        DROP TABLE IF EXISTS rincian_belanja;
        DROP TABLE IF EXISTS programs;
        DROP TABLE IF EXISTS documents;
        DROP TABLE IF EXISTS opds;
        DROP TABLE IF EXISTS summary;

        CREATE TABLE opds (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            opd TEXT NOT NULL,
            opd_type TEXT,
            UNIQUE(opd)
        );
        CREATE TABLE documents (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            opd_id INTEGER,
            filename TEXT NOT NULL,
            filepath TEXT,
            ext TEXT,
            size_bytes INTEGER,
            FOREIGN KEY(opd_id) REFERENCES opds(id)
        );
        CREATE TABLE programs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            opd_id INTEGER,
            document_id INTEGER,
            urusan TEXT,
            bidang_urusan TEXT,
            program TEXT,
            kegiatan TEXT,
            sub_kegiatan TEXT,
            sumber_dana TEXT,
            lokasi TEXT,
            total_anggaran REAL DEFAULT 0,
            FOREIGN KEY(opd_id) REFERENCES opds(id),
            FOREIGN KEY(document_id) REFERENCES documents(id)
        );
        CREATE TABLE belanja_accounts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            program_id INTEGER,
            rekening TEXT,
            uraian TEXT,
            level INTEGER,
            jumlah REAL DEFAULT 0,
            FOREIGN KEY(program_id) REFERENCES programs(id)
        );
        CREATE TABLE rincian_belanja (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            program_id INTEGER,
            uraian TEXT,
            koefisien REAL,
            satuan TEXT,
            harga REAL,
            jumlah REAL,
            FOREIGN KEY(program_id) REFERENCES programs(id)
        );
        CREATE TABLE summary (
            key TEXT PRIMARY KEY,
            value TEXT
        );
        PRAGMA synchronous = OFF;
        PRAGMA journal_mode = MEMORY;
    """)
    conn.commit()
    return conn

def classify_opd(name):
    name = name.strip().lower()
    if name.startswith('puskesmas'): return 'Puskesmas'
    if name.startswith('kecamatan') or name.startswith('kecamataan'): return 'Kecamatan'
    if 'rsud' in name or 'rumah sakit' in name: return 'RSUD'
    if name.startswith('badan'): return 'Badan'
    if 'satuan polisi' in name or 'satpol' in name: return 'Satpol PP'
    if name.startswith('sekretariat'): return 'Sekretariat'
    if name.startswith('inspektorat'): return 'Inspektorat'
    if name.startswith('dinas'): return 'Dinas'
    return 'Lainnya'

def extract_meta(rows):
    """Extract program metadata from header rows."""
    meta = {'urusan':'','bidang':'','program':'','kegiatan':'',
            'sub_kegiatan':'','sumber_dana':'','lokasi':''}
    found = set()
    
    patterns = [
        ('urusan', r'^urusan(\s+pemerintahan)?$'),
        ('bidang', r'^bidang(\s+urusan)?$'),
        ('program', r'^program$'),
        ('kegiatan', r'^kegiatan$'),
        ('sub_kegiatan', r'^sub\s*\.?\s*kegiatan$'),
        ('sumber_dana', r'^sumber(\s+pendanaan)?$'),
        ('lokasi', r'^lokasi(\s+kegiatan)?$'),
    ]
    
    for i in range(min(len(rows), 25)):
        r = rows[i]
        first = str(r[0]).strip().lower() if r[0] else ''
        
        # Pattern: label in first cell, value in subsequent cells
        for key, pat in patterns:
            if key in found: continue
            if re.match(pat, first):
                val = _get_colon_val(r)
                if val and len(val) > 3:
                    meta[key] = _clean_meta_val(key, val)
                    found.add(key)
        
        # Pattern: "Label : Value" merged in one cell
        for cell_idx, c in enumerate(r):
            s = str(c).strip() if c else ''
            if not s: continue
            s_lower = s.lower()
            for key, pat in patterns:
                if key in found: continue
                m = re.match(rf'({pat})\s*:\s*(.+)', s_lower)
                if m and m.group(2):
                    meta[key] = _clean_meta_val(key, m.group(2))
                    found.add(key)
    
    return meta

def _get_colon_val(row):
    """Get value from row after label, stripping leading colons."""
    for c in row[1:]:
        s = str(c).strip() if c else ''
        if not s or s in ('-', ': -'): continue
        s = re.sub(r'^:\s*', '', s)
        if s and s != '-': return s
    return ''

def _clean_meta_val(key, val):
    val = val.strip()
    if key in ('program','kegiatan','sub_kegiatan'):
        parts = val.split()
        if parts and re.match(r'^[\d.]+$', parts[0]):
            val = ' '.join(parts[1:]).strip()
    return val[:300]

def parse_rekening_table(rows, start_from=18):
    """Find Kode Rekening table and extract hierarchical accounts + amounts."""
    table_start = None
    table_end = None
    
    for i in range(start_from, len(rows)):
        r = rows[i]
        clean = [str(c).strip() if c else '' for c in r]
        text = ' '.join(clean).upper()
        
        if 'KODE REKENING' in text or ('KODE' in clean[0].upper() if clean else ''):
            table_start = i + 2  # skip header row(s)
            break
    
    if table_start is None:
        return [], 0
    
    accounts = []
    total_amount = 0
    
    for i in range(table_start, len(rows)):
        r = rows[i]
        vals = [str(c).strip() if c else '' for c in r]
        
        # Check if this row has a rekening code in the first column
        first = vals[0] if vals else ''
        
        # Check for total/summary row
        if any('JUMLAH' in v.upper() for v in vals):
            for c in r:
                if isinstance(c, (int, float)):
                    total_amount = max(total_amount, c)
            break
        
        # Empty row after accounts = end of table
        if not first and not any(v for v in vals):
            if accounts and len(accounts) > 2:
                break
            continue
        
        # Parse rekening code
        if re.match(r'^[\d.]+$', first):
            # Find the account name (usually second column)
            uraian = vals[1] if len(vals) > 1 else ''
            
            # Some files have the kode split (like ["5", "1", ...] vs ["5.1", ...])
            # Handle split codes: ["5", "2"] → "5.2"
            if '.' not in first and not first.startswith('5'):
                # Could be a split rekening (like 5 in col 0, 1 in col 1, etc.)
                code_parts = [first]
                for j in range(1, min(4, len(vals))):
                    if vals[j] and re.match(r'^[\d]+$', vals[j]):
                        code_parts.append(vals[j])
                        uraian = ''
                    else:
                        if not uraian:
                            uraian = vals[j]
                        break
                if len(code_parts) > 1:
                    first = '.'.join(code_parts)
            
            # Extract amount (usually in the last columns)
            amount = 0
            for c in reversed(r):
                if isinstance(c, (int, float)) and c > 0:
                    amount = c
                    break
            
            # Clean amount if it's stored as string "Rp X.XXX.XXX,XX"
            if not amount and uraian:
                for c in reversed(r):
                    s = str(c).strip() if c else ''
                    if s and ('Rp' in s or 'Rp.' in s):
                        try:
                            cleaned = s.replace('Rp', '').replace('.', '').replace(',', '.').strip()
                            amount = float(cleaned)
                        except: pass
                        if amount: break
            
            level = first.count('.') + 1
            accounts.append({
                'rekening': first,
                'uraian': uraian,
                'level': level,
                'jumlah': amount
            })
            
            if amount > total_amount:
                total_amount = amount
            
        # Check for line items (indented under a rekening)
        elif first.startswith('[') or first.startswith('-'):
            # This is a detail line item
            amount = 0
            for c in reversed(r):
                if isinstance(c, (int, float)) and c > 0:
                    amount = c
                    break
            
            if not amount:
                for c in reversed(r):
                    s = str(c).strip() if c else ''
                    if 'Rp' in s:
                        try:
                            cleaned = s.replace('Rp', '').replace('.', '').replace(',', '.').strip()
                            amount = float(cleaned)
                        except: pass
                        if amount: break
            
            desc = ' '.join(v for v in vals if v and not v.startswith('[') and not re.match(r'^[\d.]+$', v) and 'Rp' not in v)
            accounts.append({
                'rekening': '',
                'uraian': vals[0] if vals else '',
                'level': 99,
                'jumlah': amount,
                'description': desc
            })
            
            if amount > total_amount:
                total_amount = amount
    
    return accounts, total_amount

def extract_file(fpath, conn, opd_name_raw):
    opd_name = opd_name_raw.strip()
    row = conn.execute("SELECT id FROM opds WHERE opd = ?", (opd_name,))
    result = row.fetchone()
    if not result: return False
    opd_id = result[0]
    
    try:
        wb = openpyxl.load_workbook(fpath, data_only=True, read_only=True)
        filename = os.path.basename(fpath)
        size = os.path.getsize(fpath)
        
        c = conn.cursor()
        c.execute("INSERT INTO documents (opd_id, filename, filepath, ext, size_bytes) VALUES (?,?,?,?,?)",
                  (opd_id, filename, fpath, os.path.splitext(filename)[1], size))
        doc_id = c.lastrowid
        
        program_count = 0
        for sn in wb.sheetnames[:3]:
            ws = wb[sn]
            rows = list(ws.iter_rows(max_row=350, values_only=True))
            if not rows or len(rows) < 8: continue
            
            meta = extract_meta(rows)
            if not meta['program']: continue
            
            # Use Alokasi 2027 value if present (more reliable)
            total = 0
            for i, r in enumerate(rows):
                for cell in r:
                    s = str(cell).strip() if cell else ''
                    if 'ALOKASI' in s.upper() and '2027' in s.upper():
                        for c2 in r:
                            if isinstance(c2, (int, float)) and c2 > 0:
                                total = max(total, c2)
            
            accounts, table_total = parse_rekening_table(rows)
            if table_total > total:
                total = table_total
            
            c.execute("""INSERT INTO programs (opd_id, document_id, urusan, bidang_urusan, program, kegiatan, sub_kegiatan, sumber_dana, lokasi, total_anggaran)
            VALUES (?,?,?,?,?,?,?,?,?,?)""",
            (opd_id, doc_id, meta['urusan'], meta['bidang'], meta['program'],
             meta['kegiatan'], meta['sub_kegiatan'], meta['sumber_dana'], meta['lokasi'], total))
            prog_id = c.lastrowid
            program_count += 1
            
            for acct in accounts:
                if acct['rekening']:
                    c.execute("""INSERT INTO belanja_accounts (program_id, rekening, uraian, level, jumlah)
                    VALUES (?,?,?,?,?)""",
                    (prog_id, acct['rekening'], acct['uraian'], acct['level'], acct['jumlah']))
                elif acct['jumlah'] > 0:
                    desc = acct.get('description', acct['uraian'])
                    c.execute("""INSERT INTO rincian_belanja (program_id, uraian, koefisien, satuan, harga, jumlah)
                    VALUES (?,?,0,'',0,?)""", (prog_id, desc[:300], acct['jumlah']))
        
        if program_count > 0:
            conn.commit()
            wb.close()
            return True
        wb.close()
        return False
    except Exception as e:
        # print(f"EXTRACT_ERROR: {os.path.basename(fpath)}: {e}")
        return False

def main():
    print("=== Extracting Renja Kab. Bulukumba 2027 (v3 - full detail) ===")
    print(f"Source: {SRC_DIR}")
    
    conn = init_db()
    
    # Insert OPDs
    opds = set()
    for d in os.listdir(SRC_DIR):
        dpath = os.path.join(SRC_DIR, d)
        if os.path.isdir(dpath) and not d.startswith('.'):
            opds.add(d.strip())
    
    print(f"Found {len(opds)} OPDs")
    for opd in sorted(opds):
        t = classify_opd(opd)
        conn.execute("INSERT OR IGNORE INTO opds (opd, opd_type) VALUES (?,?)", (opd, t))
    conn.commit()
    
    # Walk files
    total, success = 0, 0
    for root, dirs, files in os.walk(SRC_DIR):
        opd_name = os.path.relpath(root, SRC_DIR).split(os.sep)[0].strip()
        if opd_name == '.': continue
        
        for fname in [f for f in files if f.endswith('.xlsx')]:
            fpath = os.path.join(root, fname)
            total += 1
            if extract_file(fpath, conn, opd_name):
                success += 1
            if total % 50 == 0:
                conn.commit()
                print(f"  {total}/{success}")
    
    conn.commit()
    
    counts = conn.execute("""
        SELECT (SELECT COUNT(*) FROM opds),
               (SELECT COUNT(*) FROM documents),
               (SELECT COUNT(*) FROM programs),
               (SELECT COUNT(*) FROM belanja_accounts)
    """).fetchone()
    
    print(f"\n=== Done: {total} files, {success} extracted ===")
    print(f"OPDs: {counts[0]}, Docs: {counts[1]}, Programs: {counts[2]}, Rekening: {counts[3]}")
    
    # Total anggaran per belanja type
    print("\n=== BELANJA BREAKDOWN (aggregated) ===")
    breakdown = conn.execute("""
        SELECT 
            CASE 
                WHEN rekening LIKE '5.1.01%' THEN 'Belanja Pegawai'
                WHEN rekening LIKE '5.1.02%' THEN 'Belanja Barang & Jasa'
                WHEN rekening LIKE '5.2%' THEN 'Belanja Modal'
                WHEN rekening = '5.1' THEN 'Belanja Operasi (total)'
                WHEN rekening = '5' THEN 'Total Belanja Daerah'
                ELSE 'Lainnya'
            END as kategori,
            SUM(jumlah) as total,
            COUNT(*) as items
        FROM belanja_accounts WHERE level <= 3
        GROUP BY kategori ORDER BY total DESC
    """).fetchall()
    for r in breakdown:
        print(f"  {r[0]}: Rp {r[1]:,.0f} ({r[2]} akun)")
    
    conn.execute("INSERT OR REPLACE INTO summary VALUES ('total_programs', ?)", (str(counts[2]),))
    conn.execute("INSERT OR REPLACE INTO summary VALUES ('total_akun', ?)", (str(counts[3]),))
    conn.commit()
    conn.close()

if __name__ == '__main__':
    main()

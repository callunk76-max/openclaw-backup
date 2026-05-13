#!/usr/bin/env python3
"""
Flask Dashboard for Renja Kab. Bulukumba 2027
"""
import sqlite3, os, json, urllib.parse, zipfile, tempfile, subprocess, shutil, re
import openpyxl
from flask import Flask, jsonify, render_template_string, request, redirect, flash
from pathlib import Path

UPLOAD_DIR = "/tmp/renja_uploads"
EXTRACT_DIR = "/tmp/renja_kab"
EXTRACT_SCRIPT = "/root/.openclaw/workspace/App/renja_dashboard/extract_data.py"
os.makedirs(UPLOAD_DIR, exist_ok=True)

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 200 * 1024 * 1024  # 200MB
DB_PATH = "/root/.openclaw/workspace/App/renja_dashboard/renja.db"

def get_db():
    conn = sqlite3.connect(DB_PATH, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    return conn

def json_serializable(obj):
    if isinstance(obj, (int, float)):
        if obj != obj:  # NaN check
            return 0
        return obj
    return str(obj)

# ─── TEMPLATE ────────────────────────────────────────────────────────
BASE_TEMPLATE = """
<!DOCTYPE html>
<html lang="id" data-bs-theme="dark">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Renja Dashboard - Bulukumba 2027</title>
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/css/bootstrap.min.css" rel="stylesheet">
    <link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.11.3/font/bootstrap-icons.min.css">
    <script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.1/dist/chart.umd.min.js"></script>
    <style>
        :root { --bs-primary: #0d6efd; }
        body { 
            background: #0f0f1a; 
            color: #e0e0e0;
            font-family: 'Segoe UI', system-ui, -apple-system, sans-serif;
        }
        .navbar { background: linear-gradient(135deg, #1a1a2e 0%, #16213e 100%) !important; border-bottom: 1px solid #2a2a4a; }
        .card { background: #1a1a2e; border: 1px solid #2a2a4a; border-radius: 12px; 
                transition: transform 0.2s, box-shadow 0.2s; }
        .card:hover { transform: translateY(-2px); box-shadow: 0 8px 25px rgba(0,0,0,0.3); }
        .card-header { background: #16213e; border-bottom: 1px solid #2a2a4a; font-weight: 600; }
        .stat-card { background: linear-gradient(135deg, #1a1a2e 0%, #16213e 100%); }
        .stat-card .number { font-size: 1.8rem; font-weight: 700; }
        .stat-card .label { font-size: 0.85rem; color: #888; text-transform: uppercase; letter-spacing: 0.5px; }
        .table { color: #e0e0e0; font-size: 0.9rem; }
        .table th { background: #16213e; border-color: #2a2a4a; color: #aaa; font-weight: 600; }
        .table td { border-color: #2a2a4a; vertical-align: middle; }
        .table-hover tbody tr:hover { background: #1f1f3a; }
        .chart-container { position: relative; min-height: 300px; }
        .badge-opd { font-size: 0.7rem; padding: 3px 8px; border-radius: 20px; }
        a { color: #6ea8fe; text-decoration: none; }
        a:hover { color: #9ec5fe; }
        .progress-bar-custom { height: 6px; border-radius: 3px; background: #2a2a4a; }
        .progress-bar-custom .fill { height: 100%; border-radius: 3px; background: linear-gradient(90deg, #0d6efd, #6610f2); }
        .breadcrumb { background: transparent; padding: 0; }
        .breadcrumb-item + .breadcrumb-item::before { color: #555; }
        .text-muted { color: #888 !important; }
        .btn-outline-light { border-color: #333; color: #ccc; }
        .btn-outline-light:hover { background: #2a2a4a; border-color: #444; }
        .chart-loading { text-align: center; padding: 60px; color: #666; }
        .footer { border-top: 1px solid #2a2a4a; padding: 20px 0; margin-top: 40px; color: #555; font-size: 0.85rem; }
        @media (max-width: 768px) {
            .stat-card .number { font-size: 1.3rem; }
        }
    </style>
</head>
<body>
    <nav class="navbar navbar-expand-lg navbar-dark px-3 py-2">
        <div class="container-fluid">
            <a class="navbar-brand fw-bold" href="/renja/">
                <i class="bi bi-bar-chart-fill me-2"></i>Renja Dashboard
                <span class="badge bg-primary ms-2">Bulukumba 2027</span>
            </a>
            <button class="navbar-toggler" type="button" data-bs-toggle="collapse" data-bs-target="#navbarNav">
                <span class="navbar-toggler-icon"></span>
            </button>
            <div class="collapse navbar-collapse" id="navbarNav">
                <ul class="navbar-nav ms-auto">
                    <li class="nav-item"><a class="nav-link" href="/renja/"><i class="bi bi-grid-3x3-gap-fill"></i> Overview</a></li>
                    <li class="nav-item"><a class="nav-link" href="/renja/opd"><i class="bi bi-building"></i> Per OPD</a></li>
                    <li class="nav-item"><a class="nav-link" href="/renja/anggaran"><i class="bi bi-cash-stack"></i> Anggaran</a></li>
                    <li class="nav-item"><a class="nav-link" href="/renja/belanja"><i class="bi bi-pie-chart"></i> Belanja</a></li>
                    <li class="nav-item"><a class="nav-link" href="/renja/data"><i class="bi bi-table"></i> Data</a></li>
                    <li class="nav-item"><a class="nav-link" href="/renja/upload"><i class="bi bi-cloud-upload"></i> Upload</a></li>
                </ul>
            </div>
        </div>
    </nav>
    <div class="container-fluid px-4 py-3">
        {{ content|safe }}
    </div>
    <div class="footer text-center">
        Data Renja Kabupaten Bulukumba &bull; Tahun Anggaran 2027 &bull; Pra-RKA
    </div>
    <script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/js/bootstrap.bundle.min.js"></script>
</body>
</html>
"""

# ─── API ROUTES ──────────────────────────────────────────────────────

@app.route('/renja/api/overview')
def api_overview():
    conn = get_db()
    data = conn.execute("SELECT * FROM summary").fetchall()
    d = {r['key']: r['value'] for r in data}
    
    opd_types = conn.execute("""
        SELECT opd_type, COUNT(*) as cnt 
        FROM opds GROUP BY opd_type ORDER BY cnt DESC
    """).fetchall()
    
    top_opd = conn.execute("""
        SELECT o.opd, o.opd_type, COUNT(p.id) as programs, 
               COALESCE(SUM(p.total_anggaran), 0) as total
        FROM opds o LEFT JOIN programs p ON p.opd_id = o.id
        GROUP BY o.id ORDER BY total DESC LIMIT 10
    """).fetchall()
    
    top_programs = conn.execute("""
        SELECT p.program, COALESCE(SUM(p.total_anggaran), 0) as total
        FROM programs p GROUP BY p.program
        ORDER BY total DESC LIMIT 15
    """).fetchall()
    
    conn.close()
    
    return jsonify({
        'total_opd': len(opd_types),
        'total_programs': int(d.get('total_programs', 0)),
        'total_anggaran': float(d.get('total_anggaran', 0)),
        'opd_types': [{'type': r['opd_type'], 'count': r['cnt']} for r in opd_types],
        'top_opd': [{'opd': r['opd'], 'type': r['opd_type'], 'programs': r['programs'], 'total': r['total']} for r in top_opd],
        'top_programs': [{'program': r['program'][:80], 'total': r['total']} for r in top_programs]
    })

@app.route('/renja/api/opd')
def api_opd():
    conn = get_db()
    rows = conn.execute("""
        SELECT o.opd, o.opd_type, COUNT(p.id) as programs, 
               COALESCE(SUM(p.total_anggaran), 0) as total
        FROM opds o LEFT JOIN programs p ON p.opd_id = o.id
        GROUP BY o.id ORDER BY total DESC
    """).fetchall()
    conn.close()
    return jsonify([{
        'opd': r['opd'], 'type': r['opd_type'], 
        'programs': r['programs'], 'total': r['total']
    } for r in rows])

@app.route('/renja/api/opd/<path:opd_name>')
def api_opd_detail(opd_name):
    conn = get_db()
    row = conn.execute("SELECT * FROM opds WHERE opd = ?", (opd_name,)).fetchone()
    if not row:
        conn.close()
        return jsonify({'error': 'not found'}), 404
    
    programs = conn.execute("""
        SELECT id, program, kegiatan, sub_kegiatan, sumber_dana, lokasi, total_anggaran
        FROM programs WHERE opd_id = ? ORDER BY total_anggaran DESC
    """, (row['id'],)).fetchall()
    
    conn.close()
    return jsonify({
        'opd': row['opd'],
        'type': row['opd_type'],
        'programs': [dict(r) for r in programs]
    })

@app.route('/renja/api/sumber_dana')
def api_sumber_dana():
    conn = get_db()
    rows = conn.execute("""
        SELECT COALESCE(NULLIF(sumber_dana, ''), 'Tidak Tersedia') as sumber, 
               COUNT(*) as cnt, COALESCE(SUM(total_anggaran), 0) as total
        FROM programs GROUP BY sumber ORDER BY total DESC
    """).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/renja/api/belanja')
def api_belanja():
    conn = get_db()
    rows = conn.execute("""
        SELECT 
            CASE 
                WHEN rekening LIKE '5.1.01%' THEN 'Belanja Pegawai'
                WHEN rekening LIKE '5.1.02%' THEN 'Belanja Barang & Jasa'
                WHEN rekening LIKE '5.2%' THEN 'Belanja Modal'
                ELSE 'Lainnya'
            END as kategori,
            SUM(jumlah) as total,
            COUNT(*) as items
        FROM belanja_accounts WHERE level >= 3 AND rekening NOT IN ('5.1','5.2')
        GROUP BY kategori ORDER BY total DESC
    """).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/renja/api/subrekening')
def api_subrekening():
    conn = get_db()
    rows = conn.execute("""
        SELECT 
            SUBSTR(rekening, 1, CASE 
                WHEN rekening LIKE '5.1.01%' THEN 6
                WHEN rekening LIKE '5.1.02%' THEN 6
                WHEN rekening LIKE '5.2.02%' THEN 6
                WHEN rekening LIKE '5.2.03%' THEN 6
                WHEN rekening LIKE '5.2.04%' THEN 6
                ELSE 0
            END) as subrek,
            uraian,
            SUM(jumlah) as total,
            COUNT(*) as items
        FROM belanja_accounts 
        WHERE level = 4 AND rekening != ''
        GROUP BY subrek
        ORDER BY total DESC
        LIMIT 30
    """).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/renja/api/belanja/<path:opd_name>')
def api_belanja_opd(opd_name):
    conn = get_db()
    rows = conn.execute("""
        SELECT 
            CASE 
                WHEN ba.rekening LIKE '5.1.01%' THEN 'Belanja Pegawai'
                WHEN ba.rekening LIKE '5.1.02%' THEN 'Belanja Barang & Jasa'
                WHEN ba.rekening LIKE '5.2%' THEN 'Belanja Modal'
                ELSE 'Lainnya'
            END as kategori,
            SUM(ba.jumlah) as total,
            COUNT(ba.id) as items
        FROM belanja_accounts ba
        JOIN programs p ON p.id = ba.program_id
        JOIN opds o ON o.id = p.opd_id
        WHERE o.opd = ? AND ba.level >= 3 AND ba.rekening NOT IN ('5.1','5.2')
        GROUP BY kategori ORDER BY total DESC
    """, (opd_name,)).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

# ─── PAGE ROUTES ─────────────────────────────────────────────────────

@app.route('/renja/')
def overview():
    conn = get_db()
    stats = conn.execute("SELECT * FROM summary").fetchall()
    d = {r['key']: r['value'] for r in stats}
    
    total_ang = float(d.get('total_anggaran', 0))
    total_progs = int(d.get('total_programs', 0))
    
    opd_types = conn.execute("""
        SELECT opd_type, COUNT(*) as cnt FROM opds GROUP BY opd_type ORDER BY cnt DESC
    """).fetchall()
    
    source_dana = conn.execute("""
        SELECT COALESCE(NULLIF(sumber_dana, ''), 'Tidak Ada') as sumber,
               SUM(total_anggaran) as total
        FROM programs WHERE total_anggaran > 0
        GROUP BY sumber ORDER BY total DESC
    """).fetchall()
    
    total_opds = sum(r['cnt'] for r in opd_types)
    
    conn.close()
    
    content = f"""
    <div class="row g-3 mb-4">
        <div class="col-md-3 col-6">
            <div class="card stat-card p-3 text-center h-100">
                <div class="label"><i class="bi bi-building me-1"></i>OPD</div>
                <div class="number text-primary">{total_opds}</div>
                <small class="text-muted">Satuan Kerja Perangkat Daerah</small>
            </div>
        </div>
        <div class="col-md-3 col-6">
            <div class="card stat-card p-3 text-center h-100">
                <div class="label"><i class="bi bi-folder2 me-1"></i>Program / Kegiatan</div>
                <div class="number text-success">{total_progs}</div>
                <small class="text-muted">Total entry program</small>
            </div>
        </div>
        <div class="col-md-3 col-6">
            <div class="card stat-card p-3 text-center h-100">
                <div class="label"><i class="bi bi-cash me-1"></i>Total Anggaran</div>
                <div class="number text-warning">Rp {total_ang:,.0f}</div>
                <small class="text-muted">Pra-RKA 2027 (terekstrak)</small>
            </div>
        </div>
        <div class="col-md-3 col-6">
            <div class="card stat-card p-3 text-center h-100">
                <div class="label"><i class="bi bi-file-earmark-spreadsheet me-1"></i>Dokumen</div>
                <div class="number text-info">457</div>
                <small class="text-muted">File Excel diproses</small>
            </div>
        </div>
    </div>
    
    <div class="row g-3">
        <div class="col-lg-8">
            <div class="card h-100">
                <div class="card-header d-flex justify-content-between align-items-center">
                    <span><i class="bi bi-bar-chart me-2"></i>Anggaran per OPD (Top 10)</span>
                </div>
                <div class="card-body">
                    <canvas id="chartOpd" height="300"></canvas>
                </div>
            </div>
        </div>
        <div class="col-lg-4">
            <div class="card h-100">
                <div class="card-header"><i class="bi bi-pie-chart me-2"></i>Komposisi OPD</div>
                <div class="card-body">
                    <canvas id="chartType" height="280"></canvas>
                </div>
            </div>
        </div>
    </div>
    
    <div class="row g-3 mt-2">
        <div class="col-lg-6">
            <div class="card h-100">
                <div class="card-header"><i class="bi bi-cash-stack me-2"></i>Sumber Pendanaan</div>
                <div class="card-body">
                    <canvas id="chartDana" height="250"></canvas>
                </div>
            </div>
        </div>
        <div class="col-lg-6">
            <div class="card h-100">
                <div class="card-header"><i class="bi bi-trophy me-2"></i>Program Terbesar</div>
                <div class="card-body">
                    <canvas id="chartProgram" height="250"></canvas>
                </div>
            </div>
        </div>
    </div>
    
    <script>
    async function loadCharts() {{
        const resp = await fetch('/renja/api/overview');
        const data = await resp.json();
        
        // Bar chart - Top 10 OPD
        new Chart(document.getElementById('chartOpd'), {{
            type: 'bar',
            data: {{
                labels: data.top_opd.map(r => r.opd.substring(0, 25)),
                datasets: [{{
                    label: 'Anggaran (Rp)',
                    data: data.top_opd.map(r => r.total),
                    backgroundColor: ['#0d6efd','#6610f2','#6f42c1','#d63384','#dc3545',
                                      '#fd7e14','#ffc107','#198754','#20c997','#0dcaf0'],
                    borderRadius: 4,
                }}]
            }},
            options: {{
                responsive: true, maintainAspectRatio: false,
                plugins: {{ legend: {{ display: false }} }},
                scales: {{
                    y: {{ ticks: {{ callback: v => 'Rp ' + (v/1e6).toFixed(0) + 'jt' }} }},
                    x: {{ ticks: {{ font: {{ size: 10 }} }} }}
                }}
            }}
        }});
        
        // Pie - OPD Types
        new Chart(document.getElementById('chartType'), {{
            type: 'doughnut',
            data: {{
                labels: data.opd_types.map(r => r.type),
                datasets: [{{
                    data: data.opd_types.map(r => r.count),
                    backgroundColor: ['#0d6efd','#6610f2','#20c997','#ffc107','#dc3545','#fd7e14'],
                }}]
            }},
            options: {{ responsive: true, maintainAspectRatio: false,
                plugins: {{ legend: {{ position: 'bottom', labels: {{ boxWidth: 12, padding: 12 }} }} }}
            }}
        }});
        
        // Sumber Dana - fetch from API
        const resp2 = await fetch('/renja/api/sumber_dana');
        const dana = await resp2.json();
        const topDana = dana.slice(0, 8);
        new Chart(document.getElementById('chartDana'), {{
            type: 'doughnut',
            data: {{
                labels: topDana.map(r => r.sumber.substring(0, 35)),
                datasets: [{{
                    data: topDana.map(r => r.total),
                    backgroundColor: ['#0d6efd','#6610f2','#6f42c1','#d63384','#dc3545','#fd7e14','#198754','#20c997'],
                }}]
            }},
            options: {{ responsive: true, maintainAspectRatio: false,
                plugins: {{ legend: {{ position: 'right', labels: {{ boxWidth: 12, font: {{ size: 10 }} }} }} }}
            }}
        }});
        
        // Top Programs
        new Chart(document.getElementById('chartProgram'), {{
            type: 'bar',
            data: {{
                labels: data.top_programs.map(r => r.program.substring(0, 40)),
                datasets: [{{
                    label: 'Anggaran',
                    data: data.top_programs.map(r => r.total),
                    backgroundColor: '#20c997',
                    borderRadius: 4,
                }}]
            }},
            options: {{
                responsive: true, maintainAspectRatio: false,
                indexAxis: 'y',
                plugins: {{ legend: {{ display: false }} }},
                scales: {{
                    x: {{ ticks: {{ callback: v => 'Rp ' + (v/1e6).toFixed(0) + 'jt' }} }},
                    y: {{ ticks: {{ font: {{ size: 9 }} }} }}
                }}
            }}
        }});
    }}
    loadCharts();
    </script>
    """
    return render_template_string(BASE_TEMPLATE, content=content)

@app.route('/renja/opd')
def opd_list():
    conn = get_db()
    rows = conn.execute("""
        SELECT o.opd, o.opd_type, COUNT(p.id) as programs, 
               COALESCE(SUM(p.total_anggaran), 0) as total
        FROM opds o LEFT JOIN programs p ON p.opd_id = o.id
        GROUP BY o.id ORDER BY o.opd_type, o.opd
    """).fetchall()
    conn.close()
    
    # Group by type
    grouped = {}
    for r in rows:
        t = r['opd_type'] if r['opd_type'] else 'Lainnya'
        if t not in grouped:
            grouped[t] = []
        grouped[t].append(r)
    
    cards = ''
    type_colors = {'Dinas': 'primary', 'Badan': 'info', 'Kecamatan': 'success', 
                   'Puskesmas': 'warning', 'Sekretariat': 'secondary', 
                   'Satpol PP': 'danger', 'RSUD': 'dark', 'Inspektorat': 'light'}
    
    for t in ['Dinas', 'Badan', 'Kecamatan', 'Puskesmas', 'Sekretariat', 'Satpol PP', 'RSUD', 'Inspektorat']:
        if t not in grouped:
            continue
        items = grouped[t]
        color = type_colors.get(t, 'primary')
        cards += f'<h5 class="mt-3 mb-2"><span class="badge bg-{color}">{t}</span> ({len(items)})</h5>'
        cards += '<div class="table-responsive"><table class="table table-hover align-middle"><thead><tr><th>OPD</th><th>Program</th><th class="text-end">Anggaran</th></tr></thead><tbody>'
        for r in items:
            if r['programs'] > 0:
                cards += f'<tr><td><a href="/renja/opd/{urllib.parse.quote(r["opd"], safe="")}">{r["opd"]}</a></td>'
                cards += f'<td>{r["programs"]}</td>'
                cards += f'<td class="text-end">Rp {r["total"]:,.0f}</td></tr>'
            else:
                cards += f'<tr><td>{r["opd"]}</td><td colspan="2" class="text-muted">Belum ada data</td></tr>'
        cards += '</tbody></table></div>'
    
    content = f"""
    <div class="d-flex justify-content-between align-items-center mb-3">
        <h4 class="mb-0"><i class="bi bi-building me-2"></i>Perangkat Daerah</h4>
        <span class="text-muted">{len(rows)} OPD</span>
    </div>
    {cards}
    """
    return render_template_string(BASE_TEMPLATE, content=content)

@app.route('/renja/opd/<path:opd_name>')
def opd_detail(opd_name):
    conn = get_db()
    row = conn.execute("SELECT * FROM opds WHERE opd = ?", (opd_name,)).fetchone()
    if not row:
        conn.close()
        content = f'<h4>OPD tidak ditemukan</h4><a href="/renja/opd">Kembali</a>'
        return render_template_string(BASE_TEMPLATE, content=content)
    
    programs = conn.execute("""
        SELECT p.* FROM programs p WHERE p.opd_id = ? 
        ORDER BY p.total_anggaran DESC
    """, (row['id'],)).fetchall()
    
    conn.close()
    
    total_ang = sum(p['total_anggaran'] for p in programs)
    
    table_rows = ''
    for i, p in enumerate(programs, 1):
        table_rows += f'<tr><td>{i}</td><td>{p["program"][:60]}</td>'
        table_rows += f'<td>{p["kegiatan"][:50]}</td>'
        table_rows += f'<td>{p["sub_kegiatan"][:50]}</td>'
        table_rows += f'<td>{p["sumber_dana"][:30]}</td>'
        table_rows += f'<td class="text-end">Rp {p["total_anggaran"]:,.0f}</td></tr>'
    
    content = f"""
    <nav aria-label="breadcrumb"><ol class="breadcrumb">
        <li class="breadcrumb-item"><a href="/renja/">Dashboard</a></li>
        <li class="breadcrumb-item"><a href="/renja/opd">OPD</a></li>
        <li class="breadcrumb-item active">{row['opd']}</li>
    </ol></nav>
    
    <div class="row g-3 mb-3">
        <div class="col-md-4">
            <div class="card p-3">
                <small class="text-muted">OPD</small>
                <h5 class="mb-0">{row[1] if isinstance(row, tuple) else row['opd']}</h5>
                <span class="badge bg-info mt-1" style="width:fit-content">{row['opd_type'] if not isinstance(row, tuple) else row[2]}</span>
            </div>
        </div>
        <div class="col-md-4">
            <div class="card p-3">
                <small class="text-muted">Program / Kegiatan</small>
                <h5 class="mb-0">{len(programs)}</h5>
            </div>
        </div>
        <div class="col-md-4">
            <div class="card p-3">
                <small class="text-muted">Total Anggaran</small>
                <h5 class="mb-0 text-warning">Rp {total_ang:,.0f}</h5>
    </div>
</div>
            </div>
        </div>
    </div>
    
    <div class="card">
        <div class="card-header">Daftar Program & Kegiatan</div>
        <div class="card-body p-0">
            <div class="table-responsive">
                <table class="table table-hover mb-0">
                    <thead><tr><th>#</th><th>Program</th><th>Kegiatan</th><th>Sub Kegiatan</th><th>Sumber Dana</th><th class="text-end">Anggaran</th></tr></thead>
                    <tbody>{table_rows}</tbody>
                </table>
            </div>
        </div>
    </div>
    """
    return render_template_string(BASE_TEMPLATE, content=content)

@app.route('/renja/anggaran')
def anggaran():
    content = """
    <div class="d-flex justify-content-between align-items-center mb-3">
        <h4 class="mb-0"><i class="bi bi-cash-stack me-2"></i>Analisis Anggaran</h4>
    </div>
    
    <div class="row g-3 mb-3">
        <div class="col-lg-6">
            <div class="card">
                <div class="card-header"><i class="bi bi-bar-chart me-2"></i>Anggaran per OPD</div>
                <div class="card-body">
                    <canvas id="chartAllOpd" height="400"></canvas>
                </div>
            </div>
        </div>
        <div class="col-lg-6">
            <div class="card">
                <div class="card-header"><i class="bi bi-pie-chart me-2"></i>Sumber Pendanaan</div>
                <div class="card-body">
                    <canvas id="chartSumberDana" height="400"></canvas>
                </div>
            </div>
        </div>
    </div>
    
    <div class="card">
        <div class="card-header"><i class="bi bi-trophy me-2"></i>Top 20 Program Berdasarkan Anggaran</div>
        <div class="card-body">
            <canvas id="chartTopPrograms" height="500"></canvas>
        </div>
    </div>
    
    <script>
    async function load() {
        const [r1, r2] = await Promise.all([
            fetch('/renja/api/opd').then(r => r.json()),
            fetch('/renja/api/sumber_dana').then(r => r.json())
        ]);
        
        const top20 = r1.slice(0, 20);
        new Chart(document.getElementById('chartAllOpd'), {
            type: 'bar',
            data: {
                labels: top20.map(r => r.opd.substring(0, 28)),
                datasets: [{
                    label: 'Anggaran',
                    data: top20.map(r => r.total),
                    backgroundColor: '#0d6efd',
                    borderRadius: 3,
                }]
            },
            options: {
                responsive: true, maintainAspectRatio: false,
                indexAxis: 'y',
                plugins: { legend: { display: false } },
                scales: {
                    x: { ticks: { callback: v => 'Rp ' + (v/1e6).toFixed(0) + ' jt' } },
                    y: { ticks: { font: { size: 9 } } }
                }
            }
        });
        
        const topSumber = r2.slice(0, 10);
        new Chart(document.getElementById('chartSumberDana'), {
            type: 'pie',
            data: {
                labels: topSumber.map(r => r.sumber.substring(0, 40)),
                datasets: [{
                    data: topSumber.map(r => r.total),
                    backgroundColor: ['#0d6efd','#6610f2','#6f42c1','#d63384','#dc3545',
                                      '#fd7e14','#ffc107','#198754','#20c997','#0dcaf0'],
                }]
            },
            options: {
                responsive: true, maintainAspectRatio: false,
                plugins: {
                    legend: { position: 'bottom', labels: { boxWidth: 12, font: { size: 10 } } }
                }
            }
        });
        
        // Top programs from overview
        const overview = await fetch('/renja/api/overview').then(r => r.json());
        new Chart(document.getElementById('chartTopPrograms'), {
            type: 'bar',
            data: {
                labels: overview.top_programs.map(r => r.program.substring(0, 50)),
                datasets: [{
                    label: 'Anggaran (Rp)',
                    data: overview.top_programs.map(r => r.total),
                    backgroundColor: '#20c997',
                    borderRadius: 3,
                }]
            },
            options: {
                responsive: true, maintainAspectRatio: false,
                indexAxis: 'y',
                plugins: { legend: { display: false } },
                scales: {
                    x: { ticks: { callback: v => 'Rp ' + (v/1e6).toFixed(0) + ' jt' } },
                    y: { ticks: { font: { size: 10 } } }
                }
            }
        });
    }
    load();
    </script>
    """
    return render_template_string(BASE_TEMPLATE, content=content)

@app.route('/renja/belanja')
def belanja():
    content = """
    <div class="d-flex justify-content-between align-items-center mb-3">
        <h4 class="mb-0"><i class="bi bi-pie-chart me-2"></i>Analisis Belanja Detail</h4>
    </div>
    
    <div class="row g-3 mb-3">
        <div class="col-md-4">
            <div class="card">
                <div class="card-header"><i class="bi bi-pie-chart me-2"></i>Komposisi Belanja</div>
                <div class="card-body">
                    <canvas id="chartBelanjaPie" height="280"></canvas>
                </div>
            </div>
        </div>
        <div class="col-md-8">
            <div class="card">
                <div class="card-header"><i class="bi bi-bar-chart me-2"></i>Sub Rekening (Top 20)</div>
                <div class="card-body">
                    <canvas id="chartSubrek" height="280"></canvas>
                </div>
            </div>
        </div>
    </div>
    
    <div class="card">
        <div class="card-header d-flex justify-content-between">
            <span><i class="bi bi-table me-2"></i>Rincian Belanja per OPD</span>
            <select id="filterRek" class="form-select form-select-sm" style="width:auto;" onchange="renderTable()">
                <option value="all">Semua Jenis</option>
                <option value="pegawai">Belanja Pegawai</option>
                <option value="barang">Barang & Jasa</option>
                <option value="modal">Belanja Modal</option>
            </select>
        </div>
        <div class="card-body p-0">
            <div class="table-responsive" style="max-height: 65vh;">
                <table class="table table-hover table-sm mb-0">
                    <thead class="sticky-top"><tr>
                        <th>OPD</th><th class="text-end">Pegawai</th><th class="text-end">Barang & Jasa</th>
                        <th class="text-end">Modal</th><th class="text-end">Total</th>
                    </tr></thead>
                    <tbody id="belanjaTable">
                        <tr><td colspan="5" class="text-center text-muted py-3">Memuat data...</td></tr>
                    </tbody>
                </table>
            </div>
        </div>
    </div>
    
    <script>
    let allRows = [];
    
    async function loadBelanja() {
        const [belanjaResp, opdResp, subrekResp] = await Promise.all([
            fetch('/renja/api/belanja').then(r => r.json()),
            fetch('/renja/api/opd').then(r => r.json()),
            fetch('/renja/api/subrekening').then(r => r.json())
        ]);
        
        // Pie chart - belanja types
        const pieData = belanjaResp.filter(r => r.kategori !== 'Total Belanja Daerah' && r.kategori !== 'Belanja Operasi');
        new Chart(document.getElementById('chartBelanjaPie'), {
            type: 'doughnut',
            data: {
                labels: pieData.map(r => r.kategori),
                datasets: [{
                    data: pieData.map(r => r.total),
                    backgroundColor: ['#d63384', '#0d6efd', '#20c997', '#ffc107'],
                }]
            },
            options: {
                responsive: true, maintainAspectRatio: false,
                plugins: {
                    legend: { position: 'bottom', labels: { boxWidth: 12, font: {size:11} } },
                    tooltip: { callbacks: { label: ctx => ctx.label + ': Rp ' + ctx.parsed.toLocaleString('id-ID') } }
                }
            }
        });
        
        // Bar chart - sub rekening
        new Chart(document.getElementById('chartSubrek'), {
            type: 'bar',
            data: {
                labels: subrekResp.map(r => (r.uraian || 'Lainnya').substring(0, 35)),
                datasets: [{
                    label: 'Total',
                    data: subrekResp.map(r => r.total),
                    backgroundColor: '#0d6efd',
                    borderRadius: 3,
                }]
            },
            options: {
                responsive: true, maintainAspectRatio: false,
                indexAxis: 'y',
                plugins: { legend: { display: false } },
                scales: {
                    x: { ticks: { callback: v => 'Rp ' + (v / 1e6).toFixed(0) + ' jt' } }
                }
            }
        });
        
        // Table data
        for (const opd of opdResp) {
            try {
                const resp = await fetch('/renja/api/belanja/' + encodeURIComponent(opd.opd));
                const data = await resp.json();
                allRows.push({
                    opd: opd.opd,
                    pegawai: (data.find(d => d.kategori === 'Belanja Pegawai') || {}).total || 0,
                    barang: (data.find(d => d.kategori === 'Belanja Barang & Jasa') || {}).total || 0,
                    modal: (data.find(d => d.kategori === 'Belanja Modal') || {}).total || 0
                });
            } catch(e) {}
        }
        
        renderTable();
    }
    
    function renderTable() {
        const filter = document.getElementById('filterRek').value;
        const tbody = document.getElementById('belanjaTable');
        
        let filtered = allRows.filter(r => {
            if (filter === 'pegawai') return r.pegawai > 0;
            if (filter === 'barang') return r.barang > 0;
            if (filter === 'modal') return r.modal > 0;
            return r.pegawai || r.barang || r.modal;
        });
        
        filtered.sort((a, b) => (b.pegawai + b.barang + b.modal) - (a.pegawai + a.barang + a.modal));
        
        if (filtered.length === 0) {
            tbody.innerHTML = '<tr><td colspan="5" class="text-center text-muted py-3">Tidak ada data</td></tr>';
            return;
        }
        
        tbody.innerHTML = filtered.map(r => {
            const total = r.pegawai + r.barang + r.modal;
            const opdLink = '/renja/opd/' + encodeURIComponent(r.opd);
            let bars = '';
            if (total > 0) {
                const pct = v => (v / total * 100).toFixed(0);
                bars = '<div class="progress-bar-custom mt-1"><div class="fill" style="width:' + pct(total) + '%"></div></div>';
            }
            return `<tr>
                <td><a href="${opdLink}">${r.opd.substring(0, 40)}</a></td>
                <td class="text-end">Rp ${r.pegawai.toLocaleString('id-ID')}</td>
                <td class="text-end">Rp ${r.barang.toLocaleString('id-ID')}</td>
                <td class="text-end">Rp ${r.modal.toLocaleString('id-ID')}</td>
                <td class="text-end fw-bold text-warning">Rp ${total.toLocaleString('id-ID')}</td>
            </tr>`;
        }).join('');
    }
    
    loadBelanja();
    </script>
    """
    return render_template_string(BASE_TEMPLATE, content=content)

# ─── Single xlsx extraction helpers (append mode) ────────────────────────

XLSX_META_PATTERNS = [
    ('urusan',       r'^urusan(\s+pemerintahan)?$'),
    ('bidang',       r'^bidang(\s+urusan)?$'),
    ('unit',         r'^unit(\s+organisasi)?$'),
    ('program',      r'^program$'),
    ('kegiatan',     r'^kegiatan$'),
    ('sub_kegiatan', r'^sub\s*\.?\s*kegiatan$'),
    ('sumber_dana',  r'^sumber(\s+pendanaan)?$'),
    ('lokasi',       r'^lokasi(\s+kegiatan)?$'),
]

def _get_meta_val(row, col_start=1):
    for c in row[col_start:]:
        s = str(c).strip() if c else ''
        if not s or s == '-' or s == ': -':
            continue
        s = re.sub(r'^:\s*', '', s)
        if s and s != '-':
            return s
    return ''

def _clean_opd_name(name):
    """Clean up OPD name: remove leading dash, normalize."""
    name = name.strip()
    # Remove leading dash, colon, etc
    name = re.sub(r'^[\s\-:]+', '', name)
    # Remove trailing "Kab. Bulukumba" or "Kabupaten" variations
    name = re.sub(r'\s+(Kab\.?|Kabupaten)\s+Bulukumba\s*$', '', name, flags=re.IGNORECASE).strip()
    # Remove trailing "Kab. / Kota" etc
    name = re.sub(r'\s+Kabupaten/Kota\s*$', '', name, flags=re.IGNORECASE).strip()
    return name

def _detect_opd_from_xlsx(fpath):
    """Try to detect OPD name from the Unit Organisasi field in the xlsx."""
    try:
        wb = openpyxl.load_workbook(fpath, data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(max_row=25, values_only=True))
        wb.close()
        
        for r in rows:
            first = str(r[0]).strip().lower() if r[0] else ''
            if re.match(r'^unit(\s+organisasi)?$', first):
                val = _get_meta_val(r)
                if val and ':' in val:
                    val = val.split(':', 1)[1].strip()
                # Clean: extract just the OPD name after the code
                parts = val.split()
                # Check for kode format like 1.02.0.00.0.00.01.0000
                if parts and re.match(r'^[\d.]+$', parts[0]):
                    return _clean_opd_name(' '.join(parts[1:]).strip())
                return _clean_opd_name(val.strip())
            # Also check for direct "Unit Organisasi : value" in one cell
            for c in r:
                s = str(c).strip() if c else ''
                m = re.match(r'^unit(\s+organisasi)?\s*:\s*(.+)', s, re.IGNORECASE)
                if m:
                    val = m.group(2).strip()
                    parts = val.split()
                    if parts and re.match(r'^[\d.]+$', parts[0]):
                        return _clean_opd_name(' '.join(parts[1:]).strip())
                    return _clean_opd_name(val)
    except:
        pass
    return None

def _detect_opd_from_filename(fname):
    """Try to detect OPD from the xlsx filename."""
    fname_lower = fname.lower().replace('_', ' ').replace('-', ' ')
    # Common OPD prefixes
    patterns = [
        (r'puskesmas', 'Puskesmas'),
        (r'pkm', 'Puskesmas'),
        (r'kecamatan', 'Kecamatan'),
        (r'dinas', 'Dinas'),
        (r'badan', 'Badan'),
    ]
    for pat, label in patterns:
        if re.search(pat, fname_lower):
            # Extract the full name
            return fname_lower[:50].title()
    return None

def _match_existing_opd(name, conn):
    """Try to match an OPD name against existing ones."""
    # Exact match
    cur = conn.cursor()
    cur.execute("SELECT id, opd FROM opds")
    all_opds = cur.fetchall()
    name_lower = name.lower().strip()
    
    for row in all_opds:
        if row['opd'].lower().strip() == name_lower:
            return row['opd']
    
    # Partial match (one contains the other)
    for row in all_opds:
        existing = row['opd'].lower().strip()
        if name_lower in existing or existing in name_lower:
            return row['opd']
    
    # Word match (check if significant words overlap)
    name_words = set(name_lower.split())
    for row in all_opds:
        existing_words = set(row['opd'].lower().split())
        overlap = name_words & existing_words
        if len(overlap) >= 2:
            return row['opd']
    
    return None

def process_single_xlsx(fpath, opd_name, conn):
    """Extract data from a single xlsx and insert into existing DB."""
    opd_name = opd_name.strip()
    
    # Ensure OPD exists
    cur = conn.cursor()
    cur.execute("SELECT id FROM opds WHERE opd = ?", (opd_name,))
    row = cur.fetchone()
    if row:
        opd_id = row['id']
    else:
        # Auto-classify OPD type
        opd_type = 'Dinas'
        name_lower = opd_name.lower()
        if 'puskesmas' in name_lower or 'pkm' in name_lower:
            opd_type = 'Puskesmas'
        elif 'kecamatan' in name_lower:
            opd_type = 'Kecamatan'
        elif 'badan' in name_lower:
            opd_type = 'Badan'
        elif 'rsud' in name_lower or 'rumah sakit' in name_lower:
            opd_type = 'RSUD'
        elif 'satpol' in name_lower or 'satuan polisi' in name_lower:
            opd_type = 'Satpol PP'
        elif 'sekretariat' in name_lower:
            opd_type = 'Sekretariat'
        elif 'inspektorat' in name_lower:
            opd_type = 'Inspektorat'
        
        cur.execute("INSERT INTO opds (opd, opd_type) VALUES (?,?)", (opd_name, opd_type))
        opd_id = cur.lastrowid
    
    # Record document
    filename = os.path.basename(fpath)
    size = os.path.getsize(fpath)
    cur.execute("INSERT INTO documents (opd_id, filename, filepath, ext, size_bytes) VALUES (?,?,?,?,?)",
                (opd_id, filename, fpath, '.xlsx', size))
    doc_id = cur.lastrowid
    
    # Extract metadata and line items
    try:
        wb = openpyxl.load_workbook(fpath, data_only=True, read_only=True)
    except:
        return False, "File Excel tidak bisa dibaca"
    
    program_count = 0
    for sn in wb.sheetnames[:3]:
        ws = wb[sn]
        rows = list(ws.iter_rows(max_row=250, values_only=True))
        if not rows or len(rows) < 5:
            continue
        
        # Extract metadata
        meta = {k: '' for k in ['urusan','bidang','program','kegiatan','sub_kegiatan','sumber_dana','lokasi']}
        found = set()
        for i in range(min(len(rows), 25)):
            r = rows[i]
            first = str(r[0]).strip().lower() if r[0] else ''
            if not first:
                continue
            for key, pattern in XLSX_META_PATTERNS:
                if key in found:
                    continue
                if re.match(pattern, first):
                    val = _get_meta_val(r)
                    if val and len(val) > 3:
                        if key in ('program','kegiatan','sub_kegiatan'):
                            parts = val.split()
                            if parts and re.match(r'^[\d.]+$', parts[0]):
                                val = ' '.join(parts[1:]).strip()
                        meta[key] = val[:300]
                        found.add(key)
            # Check single-cell format
            for c in r:
                s = str(c).strip() if c else ''
                if not s:
                    continue
                s_lower = s.lower()
                for key, pattern in XLSX_META_PATTERNS:
                    if key in found:
                        continue
                    m = re.match(rf'({pattern})\s*:\s*(.+)', s_lower)
                    if m and m.group(2):
                        val = m.group(2).strip()
                        if key in ('program','kegiatan','sub_kegiatan'):
                            parts = val.split()
                            if parts and re.match(r'^[\d.]+$', parts[0]):
                                val = ' '.join(parts[1:]).strip()
                        meta[key] = val[:300]
                        found.add(key)
        
        if not meta['program']:
            continue
        
        # Extract total from line items or alokasi 2027
        total = 0
        items = []
        in_table = False
        for i, r in enumerate(rows):
            if i < 25:
                for c in r:
                    s = str(c).strip() if c else ''
                    if 'ALOKASI' in s.upper() and '2027' in s.upper():
                        for c2 in r:
                            if isinstance(c2, (int, float)) and c2 > total:
                                total = c2
            
            clean_vals = [str(c).strip() if c else '' for c in r]
            text = ' '.join(clean_vals).upper()
            has_nums = [c for c in r if isinstance(c, (int, float)) and c > 0]
            
            if not in_table:
                if 'KOEFISIEN' in text and 'SATUAN' in text and 'HARGA' in text or \
                   ('BELANJA' in text and 'URAIAN' in text and 'JUMLAH' in text):
                    in_table = True
                continue
            
            if not any(clean_vals):
                if items:
                    break
                continue
            
            if any('JUMLAH' in v.upper() or 'TOTAL' in v.upper() for v in clean_vals):
                total_nums = [c for c in r if isinstance(c, (int, float)) and c > 0]
                if total_nums:
                    total = max(total_nums)
                break
            
            if has_nums:
                item = {
                    'uraian': '', 'koefisien': 0, 'satuan': '', 'harga': 0, 'jumlah': 0
                }
                descs = [v for v in clean_vals if v and len(v) > 2 and 
                        not re.match(r'^[\d.]+$', v) and not v.startswith('5.')]
                nums = sorted(has_nums)
                item['uraian'] = descs[0] if descs else ''
                item['koefisien'] = nums[0] if len(nums) >= 1 else 0
                item['harga'] = nums[-2] if len(nums) >= 3 else (nums[0] if len(nums) == 1 else 0)
                item['jumlah'] = nums[-1]
                items.append(item)
                if nums[-1] > total:
                    total = nums[-1]
        
        # Insert program
        cur.execute("""
            INSERT INTO programs (opd_id, document_id, urusan, bidang_urusan, program, kegiatan, 
                                  sub_kegiatan, sumber_dana, lokasi, total_anggaran)
            VALUES (?,?,?,?,?,?,?,?,?,?)
        """, (opd_id, doc_id, meta['urusan'], meta['bidang'], meta['program'], 
              meta['kegiatan'], meta['sub_kegiatan'], meta['sumber_dana'], meta['lokasi'], total))
        prog_id = cur.lastrowid
        program_count += 1
        
        for item in items:
            if item['uraian'] or item['jumlah'] > 0:
                cur.execute("""
                    INSERT INTO rincian_belanja (program_id, uraian, koefisien, satuan, harga, jumlah)
                    VALUES (?,?,?,?,?,?)
                """, (prog_id, item['uraian'][:300], item['koefisien'], item['satuan'][:50], item['harga'], item['jumlah']))
    
    wb.close()
    conn.commit()
    
    if program_count == 0:
        return False, "Tidak ditemukan data program dalam file"
    
    return True, f"{program_count} program berhasil ditambahkan"


@app.route('/renja/upload', methods=['GET', 'POST'], strict_slashes=False)
def upload_page():
    result_html = ""
    
    if request.method == 'POST':
        if 'zipfile' not in request.files:
            result_html = "<div class='alert alert-danger'>Tidak ada file dipilih</div>"
        else:
            file = request.files['zipfile']
            if file.filename == '':
                result_html = "<div class='alert alert-danger'>Tidak ada file dipilih</div>"
            elif file.filename.lower().endswith('.zip'):
                # ── ZIP mode: replace all data ──
                try:
                    zip_path = os.path.join(UPLOAD_DIR, file.filename)
                    file.save(zip_path)
                    
                    if os.path.exists(EXTRACT_DIR):
                        shutil.rmtree(EXTRACT_DIR)
                    
                    with zipfile.ZipFile(zip_path, 'r') as zf:
                        zf.extractall(EXTRACT_DIR)
                    
                    file_count = 0
                    xlsx_count = 0
                    for root, dirs, files in os.walk(EXTRACT_DIR):
                        file_count += len(files)
                        xlsx_count += sum(1 for f in files if f.endswith('.xlsx'))
                    
                    result = subprocess.run(
                        ['python3', EXTRACT_SCRIPT],
                        capture_output=True, text=True, timeout=300
                    )
                    
                    out_lines = result.stdout.split('\n')
                    summary_lines = [l for l in out_lines if 'OPDs:' in l or 'Total Anggaran' in l or 'Done:' in l]
                    summary_text = '<br>'.join(summary_lines)
                    
                    os.remove(zip_path)
                    
                    result_html = f"""
                    <div class="alert alert-success">
                        <h5><i class="bi bi-check-circle"></i> Upload ZIP Berhasil!</h5>
                        <p>File: <strong>{file.filename}</strong> ({file_count} file, {xlsx_count} xlsx)</p>
                        <p>{summary_text}</p>
                        <hr>
                        <a href="/renja/" class="btn btn-primary"><i class="bi bi-bar-chart-fill"></i> Lihat Dashboard</a>
                        <a href="/renja/opd" class="btn btn-outline-light ms-2">Lihat OPD</a>
                    </div>
                    """
                except subprocess.TimeoutExpired:
                    result_html = "<div class='alert alert-warning'>Proses ekstraksi terlalu lama. Silakan coba lagi.</div>"
                except Exception as e:
                    result_html = f"<div class='alert alert-danger'>Gagal: {str(e)}</div>"
            
            elif file.filename.lower().endswith('.xlsx'):
                # ── Single XLSX mode: append to existing data ──
                try:
                    xlsx_path = os.path.join(UPLOAD_DIR, file.filename)
                    file.save(xlsx_path)
                    
                    # Try to detect OPD from file
                    detected_opd = _detect_opd_from_xlsx(xlsx_path)
                    
                    if detected_opd:
                        opd_name = detected_opd
                        # Try to match with existing OPD
                        conn_tmp = sqlite3.connect(DB_PATH)
                        conn_tmp.row_factory = sqlite3.Row
                        matched = _match_existing_opd(opd_name, conn_tmp)
                        conn_tmp.close()
                        if matched:
                            opd_name = matched
                            auto_msg = f'OPD terdeteksi & dicocokkan: <strong>{opd_name}</strong>'
                        else:
                            auto_msg = f'OPD terdeteksi otomatis: <strong>{opd_name}</strong>'
                    else:
                        # Fallback: ask user to provide OPD via query param
                        opd_name = request.form.get('opd_name', '').strip()
                        if opd_name:
                            auto_msg = f'OPD: <strong>{opd_name}</strong>'
                        else:
                            opd_name = os.path.splitext(file.filename)[0].replace('_', ' ').replace('-', ' ').title()
                            auto_msg = f'OPD tidak terdeteksi, pakai nama file: <strong>{opd_name}</strong>'
                    
                    conn = sqlite3.connect(DB_PATH)
                    conn.row_factory = sqlite3.Row
                    
                    success, msg = process_single_xlsx(xlsx_path, opd_name, conn)
                    conn.close()
                    
                    os.remove(xlsx_path)
                    
                    if success:
                        result_html = f"""
                        <div class="alert alert-success">
                            <h5><i class="bi bi-check-circle"></i> Upload XLSX Berhasil!</h5>
                            <p>File: <strong>{file.filename}</strong></p>
                            <p>{auto_msg}</p>
                            <p>{msg}</p>
                            <hr>
                            <a href="/renja/" class="btn btn-primary"><i class="bi bi-bar-chart-fill"></i> Lihat Dashboard</a>
                            <a href="/renja/opd/{urllib.parse.quote(opd_name, safe='')}" class="btn btn-outline-light ms-2">Lihat OPD</a>
                        </div>
                        """
                    else:
                        result_html = f"<div class='alert alert-warning'>{msg}</div>"
                except Exception as e:
                    result_html = f"<div class='alert alert-danger'>Gagal: {str(e)}</div>"
            
            else:
                result_html = "<div class='alert alert-danger'>File harus berupa .zip atau .xlsx</div>"
    
    content = f"""
    <div class="row justify-content-center mt-3">
        <div class="col-md-8 col-lg-6">
            <div class="card">
                <div class="card-header"><i class="bi bi-cloud-upload me-2"></i>Upload Data Renja</div>
                <div class="card-body">
                    <p class="text-muted mb-3">
                        Upload file <strong>ZIP</strong> (seluruh data) atau <strong>XLSX</strong> (per OPD/program).
                    </p>
                    {result_html}
                    <form action="/renja/upload/" method="POST" enctype="multipart/form-data">
                        <div class="mb-3">
                            <label class="form-label">Pilih file <span class="text-muted">(.zip atau .xlsx)</span></label>
                            <input class="form-control" type="file" name="zipfile" accept=".zip,.xlsx" required>
                        </div>
                        <div class="mb-3" id="opdField" style="display:none;">
                            <label class="form-label">Nama OPD <span class="text-muted">(untuk file XLSX saja)</span></label>
                            <input class="form-control" type="text" name="opd_name" placeholder="Contoh: Dinas Kesehatan" id="opdInput">
                            <div class="form-text">Kosongkan jika ingin deteksi otomatis dari isi file</div>
                        </div>
                        <button type="submit" class="btn btn-primary w-100">
                            <i class="bi bi-upload me-2"></i>Upload & Proses
                        </button>
                    </form>
                </div>
            </div>
            <div class="card mt-3">
                <div class="card-header"><i class="bi bi-info-circle me-2"></i>Petunjuk</div>
                <div class="card-body small text-muted">
                    <ul class="mb-0">
                        <li><strong>ZIP</strong> &mdash; Upload file ZIP berisi folder per OPD. <span class="text-warning">Akan menghapus data lama & mengganti dengan data baru.</span></li>
                        <li><strong>XLSX</strong> &mdash; Upload file Excel per OPD/program/kegiatan. <span class="text-success">Data akan ditambahkan ke database yang sudah ada.</span></li>
                        <li>Untuk XLSX: Jika file mengandung nama OPD di metadata (Unit Organisasi), akan terdeteksi otomatis.</li>
                        <li>Proses ZIP bisa 1-3 menit, XLSX biasanya beberapa detik.</li>
                    </ul>
                </div>
            </div>
        </div>
    </div>
    <script>
    document.querySelector('input[name="zipfile"]').addEventListener('change', function() {{
        const f = this.files[0];
        const opdField = document.getElementById('opdField');
        if (f && f.name.toLowerCase().endsWith('.xlsx')) {{
            opdField.style.display = 'block';
        }} else {{
            opdField.style.display = 'none';
        }}
    }});
    </script>
    """
    return render_template_string(BASE_TEMPLATE, content=content)

@app.route('/renja/data')
def data_page():
    conn = get_db()
    rows = conn.execute("""
        SELECT o.opd as opd_name, o.opd_type, p.program, p.kegiatan, p.sub_kegiatan,
               p.sumber_dana, p.lokasi, p.total_anggaran
        FROM programs p JOIN opds o ON o.id = p.opd_id
        ORDER BY p.total_anggaran DESC LIMIT 200
    """).fetchall()
    conn.close()
    
    table = ''
    for r in rows:
        table += f'<tr><td>{r["opd_name"]}</td><td>{r["program"][:50]}</td><td>{r["kegiatan"][:40]}</td>'
        table += f'<td>{r["sumber_dana"][:25]}</td><td class="text-end">Rp {r["total_anggaran"]:,.0f}</td></tr>'
    
    content = f"""
    <div class="d-flex justify-content-between align-items-center mb-3">
        <h4 class="mb-0"><i class="bi bi-table me-2"></i>Data Program (Top 200)</h4>
    </div>
    <div class="card">
        <div class="card-body p-0">
            <div class="table-responsive" style="max-height: 70vh; overflow-y: auto;">
                <table class="table table-hover table-sm mb-0">
                    <thead class="sticky-top">
                        <tr><th>OPD</th><th>Program</th><th>Kegiatan</th><th>Sumber Dana</th><th class="text-end">Anggaran</th></tr>
                    </thead>
                    <tbody>{table}</tbody>
                </table>
            </div>
        </div>
    </div>
    """
    return render_template_string(BASE_TEMPLATE, content=content)

# ─── MAIN ────────────────────────────────────────────────────────────

if __name__ == '__main__':
    print("Starting Renja Dashboard on port 5010...")
    app.run(host='0.0.0.0', port=5010, debug=True)

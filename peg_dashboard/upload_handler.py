#!/usr/bin/env python3
"""Proses upload file ZIP/Excel/PDF → update data pegawai."""
import json, os, re, shutil, tempfile, zipfile, traceback
from openpyxl import load_workbook

DATA_DIR = "/root/.openclaw/workspace/peg_dashboard"
DATA_FILE = os.path.join(DATA_DIR, "pegawai.json")

def load_existing():
    with open(DATA_FILE) as f:
        return json.load(f)

def save_data(data):
    with open(DATA_FILE, 'w') as f:
        json.dump(data, f, indent=2, ensure_ascii=False)

def detect_skpd(filename):
    """Guess SKPD name from filename."""
    name = os.path.splitext(os.path.basename(filename))[0]
    # Clean common prefixes/suffixes
    name = re.sub(r'^(KINERJA\s+|LAPORAN\s+KINERJA\s+|LAPORAN\s+ABSEN\s+)', '', name)
    name = re.sub(r'\s+BULAN\s+\w+\s*$', '', name, flags=re.IGNORECASE)
    name = re.sub(r'\s+\w+\s*$', '', name, flags=re.IGNORECASE)  # remove "September" etc
    return name.strip()

def parse_pdf(filepath):
    """Extract pegawai from PDF using pdfplumber."""
    import pdfplumber
    employees = []
    seen_nips = set()
    
    with pdfplumber.open(filepath) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables()
            for table in tables:
                for row in table:
                    if len(row) < 4:
                        continue
                    no_cell = (row[0] or "").strip()
                    if not re.match(r'^\d+$', no_cell):
                        continue
                    
                    name_cell = row[1] or ""
                    golongan = (row[2] or "").strip()
                    jabatan = (row[3] or "").strip().replace('\n', ' ')
                    
                    # Extract NIP
                    nip_match = re.search(r'\b(\d{18})\b', name_cell)
                    nip = nip_match.group(1) if nip_match else ""
                    
                    # Extract name (before NIP)
                    name = ""
                    if nip:
                        idx = name_cell.index(nip)
                        name = name_cell[:idx].strip().rstrip(',').strip()
                    
                    if nip and nip not in seen_nips:
                        seen_nips.add(nip)
                        
                        # Parse NIP for extra info
                        tgl_lahir = f"{nip[:4]}-{nip[4:6]}-{nip[6:8]}"
                        tmt = f"{nip[8:12]}-{nip[12:14]}"
                        jk = "L" if nip[14:15] == '1' else "P"
                        
                        employees.append({
                            "nama": name,
                            "nip": nip,
                            "golongan": golongan,
                            "jabatan": jabatan,
                            "tgl_lahir": tgl_lahir,
                            "tmt": tmt,
                            "jenis_kelamin": jk
                        })
    
    return employees

def parse_excel(filepath):
    """Extract pegawai from Excel file."""
    employees = []
    seen_nips = set()
    
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    
    # Headers
    headers = {}
    for cell in ws[1]:
        if cell.value:
            h = str(cell.value).lower().strip()
            if 'nama' in h: headers['nama'] = cell.column - 1
            if 'nip' in h: headers['nip'] = cell.column - 1
            if 'gol' in h: headers['gol'] = cell.column - 1
            if 'jabatan' in h or 'jab' in h: headers['jab'] = cell.column - 1
            if 'pangkat' in h: headers['gol'] = cell.column - 1
    
    if not headers.get('nip'):
        return employees  # can't work without NIP column
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        nip_raw = str(row[headers['nip']] or "").strip()
        nip_match = re.search(r'\b(\d{15,18})\b', nip_raw)
        nip = nip_match.group(1) if nip_match else ""
        
        if not nip or nip in seen_nips:
            continue
        seen_nips.add(nip)
        
        # Pad to 18 digits if shorter
        if len(nip) == 15:
            nip = nip + "000"
        
        name = str(row[headers.get('nama', 0)] or "").strip() if 'nama' in headers else ""
        
        # Parse NIP
        tgl_lahir = ""
        tmt = ""
        jk = ""
        if len(nip) == 18 and nip.isdigit():
            tgl_lahir = f"{nip[:4]}-{nip[4:6]}-{nip[6:8]}"
            tmt = f"{nip[8:12]}-{nip[12:14]}"
            jk = "L" if nip[14:15] == '1' else "P"
        
        gol = str(row[headers.get('gol', 0)] or "").strip() if 'gol' in headers else ""
        jab = str(row[headers.get('jab', 0)] or "").strip() if 'jab' in headers else ""
        
        employees.append({
            "nama": name,
            "nip": nip,
            "golongan": gol,
            "jabatan": jab,
            "tgl_lahir": tgl_lahir,
            "tmt": tmt,
            "jenis_kelamin": jk
        })
    
    wb.close()
    return employees

def normalize_golongan(raw):
    """Quick golongan normalization."""
    if not raw or not raw.strip():
        return {"kode": "", "pangkat": "", "full": ""}
    clean = raw.replace('\n', '').strip()
    
    GOL_RE = re.compile(r"(IV[abcde]|III[a-d]|II[a-d]|I[a-d])", re.IGNORECASE)
    m = GOL_RE = re.compile(r"(IV[abcde]|III[a-d]|II[a-d]|I[a-d])", re.IGNORECASE)
    
    STANDAR = {
        "ia": ("I/a", "Juru Muda"), "ib": ("I/b", "Juru Muda Tingkat I"),
        "ic": ("I/c", "Juru"), "id": ("I/d", "Juru Tingkat I"),
        "iia": ("II/a", "Pengatur Muda"), "iib": ("II/b", "Pengatur Muda Tingkat I"),
        "iic": ("II/c", "Pengatur"), "iid": ("II/d", "Pengatur Tingkat I"),
        "iiia": ("III/a", "Penata Muda"), "iiib": ("III/b", "Penata Muda Tingkat I"),
        "iiic": ("III/c", "Penata"), "iiid": ("III/d", "Penata Tingkat I"),
        "iva": ("IV/a", "Pembina"), "ivb": ("IV/b", "Pembina Tingkat I"),
        "ivc": ("IV/c", "Pembina Muda"), "ivd": ("IV/d", "Pembina Madya"),
        "ive": ("IV/e", "Pembina Utama"),
    }
    
    if m:
        key = m.group(1).lower()
        if key in STANDAR:
            kode, pangkat = STANDAR[key]
            return {"kode": kode, "pangkat": pangkat, "full": f"{pangkat} ({kode})"}
    
    mapping = [
        ("Penata Muda Tingkat I", "III/b", "Penata Muda Tingkat I"),
        ("Penata Muda", "III/a", "Penata Muda"),
        ("Penata Tingkat I", "III/d", "Penata Tingkat I"),
        ("Penata", "III/c", "Penata"),
        ("Pembina Utama Muda", "IV/c", "Pembina Muda"),
        ("Pembina Utama Madya", "IV/d", "Pembina Madya"),
        ("Pembina Utama", "IV/e", "Pembina Utama"),
        ("Pembina Tingkat I", "IV/b", "Pembina Tingkat I"),
        ("Pembina", "IV/a", "Pembina"),
        ("Pengatur Muda Tingkat I", "II/b", "Pengatur Muda Tingkat I"),
        ("Pengatur Muda", "II/a", "Pengatur Muda"),
        ("Pengatur Tingkat I", "II/d", "Pengatur Tingkat I"),
        ("Pengatur", "II/c", "Pengatur"),
        ("Juru Muda Tingkat I", "I/b", "Juru Muda Tingkat I"),
        ("Juru Muda", "I/a", "Juru Muda"),
        ("Juru Tingkat I", "I/d", "Juru Tingkat I"),
        ("Juru", "I/c", "Juru"),
    ]
    for nama, kode, pangkat in mapping:
        if nama.lower() in clean.lower():
            return {"kode": kode, "pangkat": pangkat, "full": f"{pangkat} ({kode})"}
    return {"kode": "", "pangkat": clean, "full": clean}

def process_uploaded_files(files):
    """Process uploaded files and merge into pegawai.json."""
    existing = load_existing()
    existing_skpd = {a['skpd']: a for a in existing}
    total_new = 0
    total_updated = 0
    results = []
    
    for f in files:
        filename = f.filename
        if not filename:
            continue
        
        ext = os.path.splitext(filename)[1].lower()
        
        # Guess the SKPD from filename
        skpd_name = detect_skpd(filename)
        
        if ext == '.zip':
            # Extract and process each file in the ZIP
            with tempfile.TemporaryDirectory() as tmpdir:
                zpath = os.path.join(tmpdir, filename)
                f.save(zpath)
                with zipfile.ZipFile(zpath) as z:
                    z.extractall(tmpdir)
                
                for root, dirs, files_in_zip in os.walk(tmpdir):
                    for zf in files_in_zip:
                        zf_path = os.path.join(root, zf)
                        zf_ext = os.path.splitext(zf)[1].lower()
                        zf_skpd = detect_skpd(zf)
                        
                        if zf_ext == '.pdf':
                            emps = parse_pdf(zf_path)
                        elif zf_ext in ('.xlsx', '.xls'):
                            emps = parse_excel(zf_path)
                        else:
                            continue
                        
                        n, u = merge_employees(existing, zf_skpd, emps)
                        total_new += n
                        total_updated += u
                        results.append(f"{zf_skpd}: {len(emps)} pegawai")
        
        elif ext == '.pdf':
            fpath = os.path.join(DATA_DIR, "upload_temp.pdf")
            f.save(fpath)
            emps = parse_pdf(fpath)
            os.remove(fpath)
            n, u = merge_employees(existing, skpd_name, emps)
            total_new += n
            total_updated += u
            results.append(f"{skpd_name}: {len(emps)} pegawai")
        
        elif ext in ('.xlsx', '.xls'):
            fpath = os.path.join(DATA_DIR, "upload_temp.xlsx")
            f.save(fpath)
            emps = parse_excel(fpath)
            os.remove(fpath)
            n, u = merge_employees(existing, skpd_name, emps)
            total_new += n
            total_updated += u
            results.append(f"{skpd_name}: {len(emps)} pegawai")
    
    if total_new > 0 or total_updated > 0:
        # Re-normalize golongan
        for a in existing:
            for e in a['pegawai']:
                g = e.get('golongan', '')
                norm = normalize_golongan(g)
                e['golongan_kode'] = norm['kode']
                e['golongan_pangkat'] = norm['pangkat']
                e['golongan'] = norm['full']
        
        # Update totals
        for a in existing:
            a['total'] = len(a['pegawai'])
        
        save_data(existing)
        
        msg = f"✅ {total_new} baru, {total_updated} diperbarui"
        if results:
            msg += " — " + "; ".join(results[:5])
        return {"ok": True, "message": msg}
    
    return {"ok": False, "message": "Tidak ada data baru ditemukan"}

def merge_employees(existing_list, skpd_name, new_emps):
    """Merge new employees into existing list by SKPD."""
    # Find or create SKPD entry
    skpd_entry = None
    for a in existing_list:
        if a['skpd'].lower() == skpd_name.lower():
            skpd_entry = a
            break
    
    if not skpd_entry:
        skpd_entry = {"skpd": skpd_name, "total": 0, "pegawai": []}
        existing_list.append(skpd_entry)
    
    existing_nips = {e['nip']: e for e in skpd_entry['pegawai']}
    new_count = 0
    update_count = 0
    
    for e in new_emps:
        if not e['nip']:
            continue
        if e['nip'] in existing_nips:
            # Update existing
            old = existing_nips[e['nip']]
            for k, v in e.items():
                if v and v != old.get(k, ''):
                    old[k] = v
                    update_count += 1
        else:
            skpd_entry['pegawai'].append(e)
            new_count += 1
    
    return new_count, update_count
